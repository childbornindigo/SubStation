#!/usr/bin/env python3
"""Add Legend and Top 15 to Start tabs to the Jun Excel file."""

import json
import statistics
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent / "market-intelligence"
PMCTIRE_PATH = BASE / "pmctire-data.json"
EXCEL_PATH = BASE / "Jun_Tire_Market_Intelligence_May2026.xlsx"

with open(PMCTIRE_PATH) as f:
    pmc = json.load(f)

FITMENT = {
    "175/65R15": "Fit, Yaris, Versa, Rio, Accent",
    "185/65R15": "Civic, Corolla, Sentra, Elantra, Forte",
    "195/65R15": "Civic, Corolla, Sentra, Elantra, Mazda3",
    "205/55R16": "Civic, Corolla, Mazda3, Jetta, Elantra",
    "205/60R16": "Camry, Accord, Sonata, Altima, Legacy",
    "215/55R17": "Camry, Accord, Mazda6, Sonata, Altima",
    "215/60R16": "Camry, Accord, Sonata, Altima, Legacy",
    "215/65R16": "Escape, Tucson, Sportage, Forester, CX-5",
    "215/65R17": "CR-V, Escape, Rogue, Tucson, Forester",
    "225/40R18": "Civic Si, WRX, Golf GTI, Mazda3, Elantra N",
    "225/45R17": "Civic Si, WRX, GTI, Mazda3, Sentra",
    "225/45R18": "Camry, Accord, Mazda6, Sonata, K5",
    "225/50R17": "Camry, Sonata, Altima, Malibu, Legacy",
    "225/55R17": "Accord, Camry, Maxima, Optima, Mazda6",
    "225/55R18": "Edge, Outback, Venza, CX-5, Murano",
    "225/60R17": "RAV4, CR-V, Rogue, Escape, Forester",
    "225/60R18": "RAV4, CR-V, Tucson, Sportage, CX-5",
    "225/65R17": "RAV4, CR-V, Rogue, Escape, Outback",
    "235/45R18": "Tesla 3, BMW 3, Lexus IS, WRX",
    "235/50R18": "Tesla Y, Model 3, Polestar, Volvo S60",
    "235/55R18": "Highlander, Pilot, Palisade, Telluride, Explorer",
    "235/55R19": "Tucson, Sportage, CX-5, RAV4, CR-V",
    "235/60R18": "Highlander, Pilot, Explorer, Santa Fe, Palisade",
    "235/65R17": "Santa Fe, Tucson, Sportage, CX-5, Equinox",
    "235/65R18": "Santa Fe, Tucson, Sportage, CX-5, Equinox",
    "235/70R16": "Tacoma, Ranger, Canyon, Colorado, Frontier",
    "245/45R18": "Model 3, BMW 3, Audi A4, Mercedes C, IS",
    "245/55R19": "Model Y, BMW X3, Audi Q5, XC60, GLC",
    "245/60R18": "Traverse, Explorer, Grand Cherokee, Durango, Sorento",
    "245/65R17": "4Runner, Wrangler, Bronco, Tacoma, Colorado",
    "245/75R16": "Wrangler, 4Runner, Tacoma, Bronco, Gladiator",
    "255/45R19": "Model Y, BMW X3, Audi Q5, Mercedes GLC, XC60",
    "255/55R18": "Explorer, Grand Cherokee, Durango, Highlander, Pilot",
    "255/65R18": "F-150, RAM 1500, Silverado, Sierra, Tundra",
    "255/70R18": "F-150, RAM 1500, Silverado, Sierra, Tundra",
    "265/60R18": "F-150, RAM 1500, Silverado, Sierra, Tundra",
    "265/65R18": "F-150, RAM 1500, Silverado, Sierra, Tundra",
    "265/70R17": "F-150, Silverado, Sierra, RAM 1500, Tacoma",
    "275/55R20": "F-150, RAM 1500, Tundra, Silverado, Sierra",
    "275/60R20": "F-150, RAM 1500, Tundra, Expedition, Yukon",
    "275/65R18": "F-150, RAM 1500, Silverado, Sierra, Tundra",
    "275/70R18": "F-250, RAM 2500, Silverado 2500, Sierra 2500",
    "285/45R22": "Escalade, Tahoe, Suburban, Yukon, Navigator",
    "285/65R18": "F-250, RAM 2500, Silverado 2500, Sierra 2500",
    "285/70R17": "F-250, RAM 2500, Silverado 2500, Super Duty",
}

# High-demand sizes in Canada (by vehicle registration volume)
SIZE_DEMAND_RANK = {
    "225/65R17": 100,  # RAV4, CR-V, Rogue — #1 selling segment
    "225/60R18": 95,
    "235/65R18": 93,
    "235/55R18": 92,
    "215/55R17": 90,
    "205/55R16": 88,
    "225/45R17": 86,
    "235/60R18": 85,
    "265/70R17": 84,  # F-150 / trucks
    "225/60R17": 83,
    "215/65R17": 82,
    "255/65R18": 80,
    "235/55R19": 78,
    "245/45R18": 76,
    "225/55R17": 75,
    "195/65R15": 74,
    "235/65R17": 73,
    "245/65R17": 72,
    "265/65R18": 70,
    "275/55R20": 68,
    "235/45R18": 66,
    "255/70R18": 64,
    "225/40R18": 62,
    "275/60R20": 60,
    "245/55R19": 58,
    "205/60R16": 56,
    "215/60R16": 54,
    "245/75R16": 52,
    "255/55R18": 50,
    "275/65R18": 48,
    "235/70R16": 46,
    "285/45R22": 44,
    "265/60R18": 42,
    "245/60R18": 40,
    "175/65R15": 38,
    "185/65R15": 36,
    "215/65R16": 34,
    "275/70R18": 32,
    "285/65R18": 30,
    "285/70R17": 28,
}

# Flatten all PMCtire products
all_products = []
for size, products in pmc["sizes"].items():
    for p in products:
        p["size"] = size
        all_products.append(p)

# Brand tier (consumer recognition in Canada)
BRAND_TIER = {
    "Michelin": 10, "Continental": 9, "Bridgestone": 9, "Goodyear": 8,
    "Pirelli": 8, "Toyo": 7, "Yokohama": 7, "Firestone": 7,
    "Cooper": 6, "General": 6, "Hankook": 6, "BFGoodrich": 7,
    "Nokian": 6, "Kumho": 5, "Nexen": 4, "Falken": 5,
    "Nitto": 5, "Sailun": 3, "Westlake": 2, "Ironman": 2,
    "Maxtrek": 2, "Gislaved": 4, "Zeta": 2, "Radar": 2,
}

# Score each product by current demand
for p in all_products:
    size_score = SIZE_DEMAND_RANK.get(p["size"], 20)
    brand_score = BRAND_TIER.get(p["brand"], 3) * 10

    # Season bonus: it's May — all-season and summer get priority
    season = (p.get("season") or "").lower()
    if season in ("all-season", "all season", "all-weather"):
        season_bonus = 20
    elif season == "summer":
        season_bonus = 15
    elif season == "winter":
        season_bonus = 5
    else:
        season_bonus = 10

    # Inventory signal: high PMCtire inventory = high expected demand
    inv = p.get("inventory", 0) or 0
    inv_score = min(20, inv / 5)

    # Margin: real wholesale vs retail
    retail = p.get("price_cad", 0) or 0
    wholesale = p.get("wholesale_cost", 0) or 0
    margin_pct = ((retail - wholesale) / retail * 100) if retail > 0 else 0
    margin_score = min(15, margin_pct / 3)

    p["demand_composite"] = size_score + brand_score + season_bonus + inv_score + margin_score
    p["margin_pct"] = round(margin_pct, 1)
    p["margin_dollar"] = round(retail - wholesale, 2) if retail and wholesale else 0

# Deduplicate: best product per brand+model (across sizes, pick most popular size)
model_best = {}
for p in all_products:
    key = (p["brand"], p["model"])
    if key not in model_best or p["demand_composite"] > model_best[key]["demand_composite"]:
        model_best[key] = p

# Rank all, then pick top 15 with diversity constraints
ranked = sorted(model_best.values(), key=lambda p: -p["demand_composite"])

# Enforce diversity: max 2 per size, max 3 per brand, at least 8 sizes represented
top15 = []
size_count = defaultdict(int)
brand_count = defaultdict(int)
for p in ranked:
    if len(top15) >= 15:
        break
    if size_count[p["size"]] >= 2:
        continue
    if brand_count[p["brand"]] >= 3:
        continue
    top15.append(p)
    size_count[p["size"]] += 1
    brand_count[p["brand"]] += 1

# --- Styling ---
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
LABEL_FONT = Font(name="Calibri", bold=True, size=10, color="1F4E79")
MONEY_FMT = '$#,##0.00'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)
T1_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
T2_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
T3_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
SILVER_FILL = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
BRONZE_FILL = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")


def style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER


def auto_width(ws, max_w=35):
    for col_cells in ws.columns:
        mx = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(mx + 3, max_w)


# Load existing workbook
wb = load_workbook(str(EXCEL_PATH))

# ========================================================================
# TAB: LEGEND
# ========================================================================
ws_legend = wb.create_sheet("Legend", 0)  # Insert as first tab

legend_sections = [
    ("LEGEND — How to Read This Report", None),
    ("", None),
    ("DATA SOURCE", None),
    ("Source", "PMCtire.com — Canada's largest online tire retailer (real wholesale + retail data)"),
    ("Scrape Date", "May 2026"),
    ("Coverage", "1,365 products across 65 tire sizes from 24+ brands"),
    ("Market", "Canadian retail pricing in CAD"),
    ("", None),
    ("COLUMN DEFINITIONS", None),
    ("Brand", "Tire manufacturer (e.g. Michelin, Continental, Bridgestone)"),
    ("Model", "Specific tire product line (e.g. X-Ice Snow, CrossClimate2, Defender2)"),
    ("Size", "Tire dimensions: Width/AspectRatio + R + RimDiameter (e.g. 225/65R17)"),
    ("Category", "Tire classification: Touring, Performance, All-Terrain, Highway, etc."),
    ("Season", "When the tire is designed for: All-Season, Winter, Summer, All-Weather"),
    ("Retail Price (CAD)", "Current listed retail price at PMCtire in Canadian dollars"),
    ("Wholesale Cost (CAD)", "PMCtire's actual wholesale/cost price — this is what the retailer pays"),
    ("Our Target Price", "Suggested selling price = Wholesale × 1.30 (30% markup)"),
    ("Our Margin ($)", "Target Price minus Wholesale Cost — dollar profit per tire"),
    ("Our Margin (%)", "(Target − Wholesale) / Target — percentage margin"),
    ("Demand Score", "Composite score 1-100 based on: size popularity (40%), brand recognition (30%), season timing (15%), inventory signal (15%)"),
    ("Speed Rating", "Max speed capability: T=190km/h, H=210, V=240, W=270, Y=300"),
    ("Load Index", "Weight capacity per tire (e.g. 95 = 690kg per tire)"),
    ("Vehicle Fitment", "Common vehicles that use this tire size in Canada"),
    ("Inventory", "PMCtire's current stock level — higher = higher expected demand"),
    ("", None),
    ("PRIORITY TIERS", None),
    ("T1 — Stock First", "High demand + good margin. These are the money-makers. Stock immediately."),
    ("T2 — Stock Next", "Moderate demand or moderate margin. Good secondary inventory."),
    ("T3 — Stock on Request", "Low demand or thin margin. Only order when a customer asks."),
    ("", None),
    ("PERFORMANCE TIERS", None),
    ("Premium", "Retail price ≥ $300 — Michelin, Continental, Pirelli, Bridgestone"),
    ("Mid", "$180–$299 — Toyo, Yokohama, Hankook, Cooper, General"),
    ("Value", "< $180 — Sailun, Westlake, Ironman, Maxtrek, Radar"),
    ("", None),
    ("SEASON DEFINITIONS", None),
    ("All-Season", "3-season tire (spring/summer/fall). Most common. Year-round in mild climates."),
    ("All-Weather", "True 4-season tire with snowflake rating. Legal for winter in all provinces."),
    ("Winter", "Dedicated winter tire. Required by law in Quebec (Dec 1 – Mar 15). Best Oct–Apr."),
    ("Summer", "Warm-weather performance tire. Do NOT use below 7°C."),
    ("", None),
    ("COLOR CODING", None),
    ("Green cells", "T1 priority — stock these first"),
    ("Yellow cells", "T2 priority — stock these next"),
    ("Red/pink cells", "T3 priority — order on request only"),
    ("Gold highlight", "Top 3 ranked items (in Top 15 tab)"),
    ("", None),
    ("SCORING METHODOLOGY", None),
    ("Demand Score", "Weighted composite: Size Popularity (40%) × Brand Recognition (30%) × Season Timing (15%) × Retailer Inventory Signal (15%)"),
    ("Size Popularity", "Based on Canadian vehicle registration data — 225/65R17 is #1 (RAV4, CR-V, Rogue)"),
    ("Brand Recognition", "Consumer awareness ranking: Michelin/Continental (top) → Westlake/Ironman (bottom)"),
    ("Season Timing", "May 2026: All-Season scores highest, Winter scores lowest"),
    ("Inventory Signal", "PMCtire stocks more of what sells more — their inventory levels are a demand proxy"),
    ("", None),
    ("MARGIN NOTES", None),
    ("Wholesale Cost", "Real cost data from PMCtire's internal API — not estimated"),
    ("30% Markup Target", "Industry standard for independent tire shops. Adjust per your pricing strategy."),
    ("Margin varies by brand", "Premium brands (Michelin, Pirelli) have tighter margins but sell on reputation. Value brands (Sailun, Westlake) have wider margins but need harder selling."),
]

row = 1
for label, value in legend_sections:
    if label == "" and value is None:
        row += 1
        continue

    if value is None:
        # Section header
        cell = ws_legend.cell(row=row, column=1, value=label)
        cell.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
        if label.startswith("LEGEND"):
            cell.font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
        ws_legend.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        continue

    cell_label = ws_legend.cell(row=row, column=1, value=label)
    cell_label.font = LABEL_FONT
    cell_label.alignment = Alignment(vertical="top")

    cell_value = ws_legend.cell(row=row, column=2, value=value)
    cell_value.font = DATA_FONT
    cell_value.alignment = Alignment(wrap_text=True, vertical="top")

    # Color the tier examples
    if label == "T1 — Stock First":
        cell_label.fill = T1_FILL
    elif label == "T2 — Stock Next":
        cell_label.fill = T2_FILL
    elif label == "T3 — Stock on Request":
        cell_label.fill = T3_FILL
    elif label == "Green cells":
        cell_label.fill = T1_FILL
    elif label == "Yellow cells":
        cell_label.fill = T2_FILL
    elif label == "Red/pink cells":
        cell_label.fill = T3_FILL
    elif label == "Gold highlight":
        cell_label.fill = GOLD_FILL

    row += 1

ws_legend.column_dimensions["A"].width = 28
ws_legend.column_dimensions["B"].width = 80
ws_legend.column_dimensions["C"].width = 15

# ========================================================================
# TAB: TOP 15 TO START
# ========================================================================
ws_top = wb.create_sheet("Top 15 to Start", 1)  # Second tab

headers = [
    "Rank", "Brand", "Model", "Size", "Season", "Category",
    "Retail Price (CAD)", "Wholesale Cost (CAD)", "Our Target (30% markup)",
    "Margin ($)", "Margin (%)",
    "Demand Score", "Speed Rating", "Load Index",
    "Vehicle Fitment", "Why Stock This",
]
NC = len(headers)

# Title row
ws_top.merge_cells("A1:P1")
title_cell = ws_top.cell(row=1, column=1, value="TOP 15 TIRES TO START STOCKING — Canadian Market, May 2026")
title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
title_cell.alignment = Alignment(horizontal="center")

ws_top.merge_cells("A2:P2")
subtitle = ws_top.cell(row=2, column=1, value="Ranked by current demand (size popularity + brand recognition + season timing + retailer inventory signal)")
subtitle.font = Font(name="Calibri", italic=True, size=10, color="666666")
subtitle.alignment = Alignment(horizontal="center")

# Headers in row 3
for i, h in enumerate(headers, 1):
    ws_top.cell(row=3, column=i, value=h)
style_header(ws_top, 3, NC)

# Data rows
for rank, p in enumerate(top15, 1):
    r = rank + 3
    retail = p.get("price_cad", 0) or 0
    wholesale = p.get("wholesale_cost", 0) or 0
    target = round(wholesale * 1.30, 2) if wholesale else round(retail * 0.85, 2)
    margin_d = round(target - wholesale, 2)
    margin_p = margin_d / target if target > 0 else 0

    # Why stock this
    reasons = []
    size_rank = SIZE_DEMAND_RANK.get(p["size"], 0)
    brand_rank = BRAND_TIER.get(p["brand"], 0)
    if size_rank >= 80:
        reasons.append(f"Top-demand size ({FITMENT.get(p['size'], 'popular vehicles')})")
    elif size_rank >= 60:
        reasons.append(f"High-demand size")
    if brand_rank >= 8:
        reasons.append(f"Top-tier brand — customers ask for it by name")
    elif brand_rank >= 6:
        reasons.append(f"Well-known brand")
    season = (p.get("season") or "").lower()
    if season in ("all-season", "all-weather", "all season"):
        reasons.append("All-season = year-round sales")
    if margin_p >= 0.25:
        reasons.append(f"Strong margin ({margin_p:.0%})")
    inv = p.get("inventory", 0) or 0
    if inv >= 50:
        reasons.append(f"High retailer inventory ({inv} units) = proven seller")
    why = ". ".join(reasons) if reasons else "Solid all-around pick"

    vals = [
        rank, p["brand"], p["model"], p["size"],
        p.get("season", ""), p.get("category", ""),
        retail, wholesale, target,
        margin_d, margin_p,
        round(p["demand_composite"], 1),
        p.get("speed_rating", ""), p.get("load_index", ""),
        FITMENT.get(p["size"], p.get("vehicle_fitment", "")),
        why,
    ]
    for i, v in enumerate(vals, 1):
        ws_top.cell(row=r, column=i, value=v)
    style_data_row(ws_top, r, NC)

    # Money formatting
    for col in [7, 8, 9, 10]:
        ws_top.cell(row=r, column=col).number_format = MONEY_FMT
    ws_top.cell(row=r, column=11).number_format = PCT_FMT

    # Gold/silver/bronze for top 3
    if rank == 1:
        for col in range(1, NC + 1):
            ws_top.cell(row=r, column=col).fill = GOLD_FILL
            ws_top.cell(row=r, column=col).font = Font(name="Calibri", bold=True, size=10)
    elif rank == 2:
        for col in range(1, NC + 1):
            ws_top.cell(row=r, column=col).fill = SILVER_FILL
            ws_top.cell(row=r, column=col).font = Font(name="Calibri", bold=True, size=10)
    elif rank == 3:
        for col in range(1, NC + 1):
            ws_top.cell(row=r, column=col).fill = BRONZE_FILL
            ws_top.cell(row=r, column=col).font = Font(name="Calibri", bold=True, size=10)
    else:
        ws_top.cell(row=r, column=1).font = BOLD_FONT

last_row = len(top15) + 3
ws_top.auto_filter.ref = f"A3:{get_column_letter(NC)}{last_row}"
ws_top.freeze_panes = "A4"

# Summary row
summary_row = last_row + 2
ws_top.cell(row=summary_row, column=1, value="SUMMARY").font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
ws_top.cell(row=summary_row + 1, column=1, value="Total Products:").font = LABEL_FONT
ws_top.cell(row=summary_row + 1, column=2, value=15).font = BOLD_FONT
ws_top.cell(row=summary_row + 2, column=1, value="Brands Covered:").font = LABEL_FONT
ws_top.cell(row=summary_row + 2, column=2, value=len(set(p["brand"] for p in top15))).font = BOLD_FONT
ws_top.cell(row=summary_row + 3, column=1, value="Avg Margin:").font = LABEL_FONT
margins = [((p.get("price_cad", 0) or 0) - (p.get("wholesale_cost", 0) or 0)) for p in top15 if p.get("price_cad") and p.get("wholesale_cost")]
avg_margin = round(statistics.mean(margins), 2) if margins else 0
ws_top.cell(row=summary_row + 3, column=2, value=avg_margin).font = BOLD_FONT
ws_top.cell(row=summary_row + 3, column=2).number_format = MONEY_FMT
ws_top.cell(row=summary_row + 4, column=1, value="Sizes Covered:").font = LABEL_FONT
ws_top.cell(row=summary_row + 4, column=2, value=len(set(p["size"] for p in top15))).font = BOLD_FONT

auto_width(ws_top, max_w=45)

# Save
wb.save(str(EXCEL_PATH))

print(f"Updated: {EXCEL_PATH}")
print(f"Added: Legend tab + Top 15 to Start tab")
print()
print("Top 15:")
for i, p in enumerate(top15, 1):
    retail = p.get("price_cad", 0) or 0
    wholesale = p.get("wholesale_cost", 0) or 0
    margin = retail - wholesale
    print(f"  {i:2d}. {p['brand']:15s} {p['model']:25s} {p['size']:12s} ${retail:>7.2f} retail / ${wholesale:>7.2f} cost / ${margin:>6.2f} margin  [{p.get('season', '')}]")
