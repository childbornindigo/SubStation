#!/usr/bin/env python3
"""
build_excel.py — Build professional tire market intelligence Excel for Jun.

Merges:
  1. Google Sheet curated data (100 products, 20 sizes) — gold standard
  2. PMCtire API data (full specs, pricing, wholesale cost) — primary source

Output: market-intelligence/Jun_Tire_Market_Intelligence_May2026.xlsx
"""

import csv
import json
import os
import sys
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MI_DIR = os.path.join(os.path.dirname(SCRIPT_DIR), "market-intelligence")

GSHEET_PATH = "/tmp/jun-catalogue.csv"
PMCTIRE_PATH = os.path.join(MI_DIR, "pmctire-data.json")

FITMENT = {
    "175/65R15": "Honda Fit, Toyota Yaris, Nissan Versa",
    "185/60R15": "Honda Civic, Toyota Corolla, Mazda3",
    "185/65R15": "Toyota Corolla, Honda Civic, Hyundai Elantra",
    "195/50R16": "Mini Cooper, Honda Fit Sport",
    "195/55R16": "VW Golf, Honda Civic, Toyota Corolla",
    "195/60R15": "Honda Civic, Toyota Corolla, Nissan Sentra",
    "195/65R15": "Toyota Corolla, Honda Civic, VW Jetta, Mazda3",
    "205/45R17": "Honda Civic Si, VW GTI, Mazda3 Sport",
    "205/50R17": "Honda Accord, Toyota Camry, Mazda6",
    "205/55R16": "Honda Civic, Toyota Corolla, VW Jetta, Mazda3",
    "205/60R16": "Honda Accord, Toyota Camry, Nissan Altima",
    "205/65R16": "Honda CR-V, Toyota RAV4, Subaru Forester",
    "215/45R17": "Honda Civic Si, Mazda3, VW GTI, Subaru WRX",
    "215/50R17": "Honda Accord, Mazda6, Subaru Legacy",
    "215/55R17": "Toyota Camry, Honda Accord, Hyundai Sonata",
    "215/60R16": "Toyota Camry, Honda Accord, Hyundai Sonata",
    "215/65R17": "Honda CR-V, Toyota RAV4, Ford Escape, Hyundai Tucson",
    "215/70R16": "Honda CR-V, Toyota RAV4, Subaru Outback",
    "225/40R18": "BMW 3 Series, Audi A4, Mercedes C-Class",
    "225/45R17": "Honda Accord, Toyota Camry, BMW 3 Series",
    "225/45R18": "BMW 3 Series, Audi A4, Mercedes C-Class",
    "225/50R17": "Honda Accord, Toyota Camry, Subaru Outback",
    "225/50R18": "BMW X1, Audi Q3, Mercedes GLA",
    "225/55R17": "Honda Accord, Toyota Camry, Subaru Outback",
    "225/60R17": "Toyota RAV4, Honda CR-V, Nissan Rogue",
    "225/60R18": "Toyota Highlander, Honda Pilot, Ford Edge",
    "225/65R17": "Toyota RAV4, Honda CR-V, Ford Escape, Nissan Rogue",
    "235/40R19": "BMW 4 Series, Audi S4, Mercedes C43 AMG",
    "235/45R18": "BMW 3 Series, Audi A4, Lexus IS",
    "235/55R18": "Acura RDX, Lincoln MKC, BMW X3",
    "235/55R19": "Acura RDX, BMW X3, Mercedes GLC",
    "235/60R18": "Toyota Highlander, Honda Pilot, Ford Explorer",
    "235/65R18": "Toyota Highlander, Chevy Traverse, Ford Explorer",
    "235/70R16": "Toyota 4Runner, Jeep Wrangler, Ford Ranger",
    "245/35R20": "BMW 5 Series, Audi A6, Mercedes E-Class",
    "245/40R18": "BMW 3 Series, Audi A4, Mercedes C-Class",
    "245/45R18": "BMW 5 Series, Audi A6, Mercedes E-Class",
    "245/55R19": "Ford Explorer, Chevy Blazer, Honda Passport",
    "245/60R18": "Toyota Highlander, Chevy Traverse, GMC Acadia",
    "245/65R17": "Toyota 4Runner, Jeep Grand Cherokee, Ford Explorer",
    "245/75R16": "Toyota Tacoma, Jeep Wrangler, Ford F-150",
    "255/35R19": "BMW M3, Audi RS4, Mercedes AMG C63",
    "255/45R19": "BMW X3, Audi Q5, Mercedes GLC",
    "255/50R19": "BMW X5, Audi Q7, Mercedes GLE",
    "255/55R18": "Jeep Grand Cherokee, Ford Explorer, Dodge Durango",
    "255/65R18": "Ford F-150, RAM 1500, Chevy Silverado",
    "255/70R18": "Ford F-150, RAM 1500, Toyota Tundra",
    "265/50R20": "Chevy Tahoe, GMC Yukon, Ford Expedition",
    "265/60R18": "Toyota Tundra, Nissan Titan, Ford F-150",
    "265/60R20": "Chevy Tahoe, GMC Yukon, Ford Expedition",
    "265/65R18": "Toyota Tundra, Nissan Titan, Ford F-150",
    "265/70R16": "Toyota Tacoma, Jeep Wrangler, Ford Ranger",
    "265/70R17": "Toyota 4Runner, Jeep Wrangler, Ford F-150",
    "265/75R16": "Ford F-250, Chevy 2500, RAM 2500",
    "275/40R20": "BMW X5, Audi Q7, Porsche Cayenne",
    "275/45R20": "Chevy Tahoe, GMC Yukon, Cadillac Escalade",
    "275/55R19": "Ford Explorer, Jeep Grand Cherokee, GMC Acadia",
    "275/55R20": "Ford F-150, Chevy Silverado, RAM 1500",
    "275/60R20": "Ford F-150, Chevy Silverado, RAM 1500",
    "275/65R18": "Ford F-150, Chevy Silverado, RAM 1500, Toyota Tundra",
    "275/70R18": "Ford F-250, Chevy 2500, RAM 2500",
    "285/45R22": "Cadillac Escalade, Chevy Tahoe, GMC Yukon",
    "285/65R18": "Ford F-250, Chevy 2500, RAM 2500",
    "285/70R17": "Ford F-250, Chevy 2500, RAM 2500",
    "285/75R16": "Ford F-250, Chevy 2500, RAM 2500",
}

HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
SUB_FONT = Font(bold=True, size=11, color="1F4E79")
SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
T1_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
T2_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
T3_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
SILVER = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
BRONZE = PatternFill(start_color="DEB887", end_color="DEB887", fill_type="solid")
MEDALS = [GOLD, SILVER, BRONZE]
TIER_MAP = {"T1": T1_FILL, "T2": T2_FILL, "T3": T3_FILL}
THIN = Border(
    left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
)
CURR = "#,##0.00"
PCT = "0.0%"
VERIFIED_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def sf(v):
    if not v:
        return None
    try:
        return float(str(v).replace("$", "").replace(",", "").replace("%", "").strip())
    except ValueError:
        return None


def si(v):
    if not v:
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return None


def hdr_row(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN


def style_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).border = THIN


def normalize_season(usage):
    if not usage:
        return "All-Season"
    u = usage.lower()
    if "winter" in u:
        return "Winter"
    if "4 season" in u or "all weather" in u or "all-weather" in u:
        return "All-Weather"
    if "summer" in u:
        return "Summer / All-Season"
    return "All-Season"


def season_push(season):
    if season == "Winter":
        return "Sep-Nov push"
    if "Summer" in season:
        return "Mar-May push"
    return "Year-round (spring + fall peaks)"


def perf_tier(price):
    if not price:
        return ""
    if price < 150:
        return "Value"
    if price < 250:
        return "Mid"
    return "Premium"


def priority_tier(price, season, size):
    popular = {
        "225/65R17", "215/65R17", "205/55R16", "195/65R15",
        "265/70R17", "225/60R18", "235/65R18", "245/75R16",
    }
    if size in popular:
        return "T1"
    if price and price >= 200:
        return "T2"
    return "T3"


def features_str(p):
    feats = []
    if p.get("runflat"):
        feats.append("Runflat")
    if p.get("studdable"):
        feats.append("Studdable")
    if p.get("studded"):
        feats.append("Factory Studded")
    if p.get("ev_optimized"):
        feats.append("EV Optimized")
    return ", ".join(feats) if feats else ""


def normalize_brand(brand):
    return {"Cooper Tires": "Cooper", "General Tire": "General", "Toyo Tires": "Toyo"}.get(brand, brand)


def load_data():
    gsheet = []
    if os.path.exists(GSHEET_PATH):
        with open(GSHEET_PATH) as f:
            gsheet = list(csv.DictReader(f))
        print(f"Google Sheet: {len(gsheet)} curated products")

    pmctire = {"sizes": {}}
    if os.path.exists(PMCTIRE_PATH):
        with open(PMCTIRE_PATH) as f:
            pmctire = json.load(f)
        total = sum(len(v) for v in pmctire["sizes"].values())
        print(f"PMCtire: {total} products across {len(pmctire['sizes'])} sizes")

    return gsheet, pmctire


def build_master_catalogue(wb, gsheet, pmctire):
    ws = wb.active
    ws.title = "Master Catalogue"
    ws.sheet_properties.tabColor = "1F4E79"

    ws.merge_cells("A1:R1")
    ws["A1"] = "TIRE MASTER CATALOGUE \u2014 Canadian Market Intelligence"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:R2")
    pmc_total = sum(len(v) for v in pmctire["sizes"].values())
    ws["A2"] = f"{len(gsheet)} verified + {pmc_total} PMCtire products | May 2026 | All prices CAD"
    ws["A2"].font = Font(italic=True, color="666666")

    headers = [
        "Brand", "Model", "Size", "Category", "Season", "Speed Rating",
        "Load Index", "Load Range", "Retail Price", "Wholesale Cost",
        "Margin (%)", "Performance Tier", "Vehicle Fitment", "Features",
        "Priority", "Season Push", "Inventory", "Data Source",
    ]
    hdr_row(ws, 4, headers)
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}5000"

    money_cols = {9, 10}
    r = 5

    for row in gsheet:
        avg_price = sf(row.get("Avg Retail Price"))
        wholesale = sf(row.get("Est. Wholesale (65%)"))
        margin_pct = None
        if avg_price and wholesale and avg_price > 0:
            margin_pct = (avg_price - wholesale) / avg_price

        vals = [
            row.get("Brand", ""), row.get("Model", ""), row.get("Size", ""),
            row.get("Category", ""),
            "Winter" if "Sep" in row.get("Season Push", "") else ("Summer" if "Mar" in row.get("Season Push", "") else "All-Season"),
            row.get("Speed Rating", ""), row.get("Load Index", ""), "",
            avg_price, wholesale, margin_pct,
            row.get("Performance Tier", ""), row.get("Top Vehicle Fitment", ""), "",
            row.get("Priority Tier", ""), row.get("Season Push", ""), "In Stock", "Verified (Curated)",
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            if ci in money_cols and v is not None:
                cell.number_format = CURR
            elif ci == 11 and v is not None:
                cell.number_format = PCT
            cell.border = THIN
            cell.fill = VERIFIED_FILL
        tier = row.get("Priority Tier", "")
        if tier in TIER_MAP:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=r, column=ci).fill = TIER_MAP[tier]
        r += 1

    verified_count = r - 5

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
    sep = ws.cell(row=r, column=1, value=f"\u2500\u2500 PMCtire DATA ({pmc_total} products \u00b7 full specs \u00b7 wholesale pricing) \u2500\u2500")
    sep.font = Font(bold=True, size=11, color="1F4E79")
    sep.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    r += 1

    gsheet_keys = set((row["Brand"], row["Model"], row["Size"]) for row in gsheet)
    pmc_count = 0
    for size in sorted(pmctire["sizes"].keys()):
        for p in pmctire["sizes"][size]:
            brand = normalize_brand(p["brand"])
            model = p["model"]
            if (brand, model, size) in gsheet_keys:
                continue
            price = p.get("price_cad")
            cost = p.get("wholesale_cost")
            margin_pct = (price - cost) / price if price and cost and price > 0 else None
            season = normalize_season(p.get("season"))
            tier = priority_tier(price, season, size)
            inv = p.get("inventory")
            inv_str = f"{inv} in stock" if inv and inv > 0 else ("Low stock" if inv == 0 else "")

            vals = [
                brand, model, size, p.get("category", ""), season,
                p.get("speed_rating", ""), p.get("load_index", ""), p.get("load_range", ""),
                price, cost, margin_pct, perf_tier(price),
                FITMENT.get(size, ""), features_str(p), tier, season_push(season), inv_str, "PMCtire API",
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=ci, value=v)
                if ci in money_cols and v is not None:
                    cell.number_format = CURR
                elif ci == 11 and v is not None:
                    cell.number_format = PCT
                cell.border = THIN
            if tier in TIER_MAP:
                for ci in range(1, len(headers) + 1):
                    ws.cell(row=r, column=ci).fill = TIER_MAP[tier]
            r += 1
            pmc_count += 1

    widths = [14, 28, 12, 16, 16, 8, 8, 10, 12, 12, 10, 12, 40, 22, 7, 28, 14, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "D5"
    print(f"Master Catalogue: {verified_count} verified + {pmc_count} PMCtire = {verified_count + pmc_count} total")


def build_price_intelligence(wb, pmctire, gsheet):
    ws = wb.create_sheet("Price Intelligence by Size")
    ws.sheet_properties.tabColor = "2E75B6"
    ws.merge_cells("A1:L1")
    ws["A1"] = "PRICE INTELLIGENCE BY SIZE"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:L2")
    ws["A2"] = "Retail prices and wholesale costs from PMCtire.com | CAD pricing"
    ws["A2"].font = Font(italic=True, color="666666")

    headers = [
        "Size", "Vehicle Fitment", "# Models", "Min Price", "Avg Price",
        "Max Price", "Avg Cost", "Avg Margin $", "Margin %",
        "Best Value", "Premium Pick", "Data Quality",
    ]
    hdr_row(ws, 4, headers)
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}200"
    gsheet_sizes = set(row["Size"] for row in gsheet)

    r = 5
    for size in sorted(pmctire["sizes"].keys()):
        products = pmctire["sizes"][size]
        if not products:
            continue
        prices = [p["price_cad"] for p in products if p.get("price_cad")]
        costs = [p["wholesale_cost"] for p in products if p.get("wholesale_cost")]
        margins = []
        for p in products:
            pr, co = p.get("price_cad"), p.get("wholesale_cost")
            if pr and co and pr > 0:
                margins.append((pr - co) / pr)
        if not prices:
            continue
        cheapest = min(products, key=lambda x: x.get("price_cad") or 99999)
        priciest = max(products, key=lambda x: x.get("price_cad") or 0)
        avg_margin_dollar = sum(p["price_cad"] - p["wholesale_cost"] for p in products if p.get("price_cad") and p.get("wholesale_cost")) / len(costs) if costs else None
        quality = "Verified + PMCtire" if size in gsheet_sizes else "PMCtire"

        vals = [
            size, FITMENT.get(size, ""), len(products),
            min(prices), sum(prices) / len(prices), max(prices),
            sum(costs) / len(costs) if costs else None,
            avg_margin_dollar,
            sum(margins) / len(margins) if margins else None,
            f"{normalize_brand(cheapest['brand'])} {cheapest['model']}",
            f"{normalize_brand(priciest['brand'])} {priciest['model']}",
            quality,
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            if ci in {4, 5, 6, 7, 8}:
                cell.number_format = CURR
            elif ci == 9 and v is not None:
                cell.number_format = PCT
            cell.border = THIN
        if size in gsheet_sizes:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=r, column=ci).fill = VERIFIED_FILL
        r += 1

    widths = [12, 42, 10, 10, 10, 10, 10, 10, 10, 28, 28, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "C5"
    print(f"Price Intelligence: {r - 5} sizes")


def build_seasonal_calendar(wb, pmctire):
    ws = wb.create_sheet("Seasonal Calendar")
    ws.sheet_properties.tabColor = "548235"
    ws.merge_cells("A1:N1")
    ws["A1"] = "SEASONAL PUSH CALENDAR"
    ws["A1"].font = TITLE_FONT
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    r = 3
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
    ws.cell(row=r, column=1, value="MONTHLY INTENSITY GUIDE").font = SUB_FONT
    r += 1
    hdr_row(ws, r, ["Category", "Focus"] + months)
    r += 1

    seasons = [
        ("All-Season / Touring", "Year-round (peaks spring + fall)", [1, 1, 2, 3, 3, 1, 1, 1, 2, 3, 3, 1]),
        ("All-Terrain / Highway", "Year-round (spring peak)", [1, 1, 2, 3, 3, 2, 1, 1, 1, 2, 2, 1]),
        ("Summer / Performance", "Mar-May push", [0, 1, 3, 3, 3, 2, 2, 2, 1, 0, 0, 0]),
        ("Winter / Ice", "Sep-Nov push", [0, 0, 0, 0, 0, 0, 0, 1, 3, 3, 3, 2]),
        ("All-Weather", "Year-round (peaks at season change)", [1, 2, 3, 3, 2, 1, 1, 1, 2, 3, 3, 1]),
    ]
    fills = {
        0: PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        1: PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        2: PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid"),
        3: PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
    }
    labels = {0: "", 1: "Low", 2: "Med", 3: "HIGH"}
    for cat, focus, intensities in seasons:
        ws.cell(row=r, column=1, value=cat).font = Font(bold=True)
        ws.cell(row=r, column=2, value=focus)
        for mi, intensity in enumerate(intensities):
            cell = ws.cell(row=r, column=3 + mi, value=labels[intensity])
            cell.fill = fills[intensity]
            cell.alignment = Alignment(horizontal="center")
            if intensity == 3:
                cell.font = Font(bold=True, color="FFFFFF")
            cell.border = THIN
        style_row(ws, r, 14)
        r += 1

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
    ws.cell(row=r, column=1, value="KEY ACTION WINDOWS").font = SUB_FONT
    r += 1
    actions = [
        ("Feb-Mar", "Stock up all-season + all-terrain ahead of spring rush. Target RAV4/CR-V/Rogue owners."),
        ("Mar-May", "Push summer/performance tires \u2014 BMW, Porsche, sports car owners. Pilot Sport, P Zero."),
        ("Apr-May", "PEAK all-season demand. RAV4, CR-V, F-150 owners switching from winter. Run T1 promotions."),
        ("Aug-Sep", "Start winter tire marketing. Early-bird pricing captures prepared buyers."),
        ("Sep-Nov", "PEAK winter push. X-Ice Snow, VikingContact 8, Blizzak WS90. Stock 225/65R17, 265/70R17."),
        ("Oct-Nov", "Second all-season peak. Last chance before winter-only demand."),
        ("Dec-Jan", "Low season. Focus on online presence, reviews, SEO. Clearance on remaining summer stock."),
    ]
    for period, action in actions:
        ws.cell(row=r, column=1, value=period).font = Font(bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=14)
        ws.cell(row=r, column=2, value=action)
        style_row(ws, r, 14)
        r += 1

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
    ws.cell(row=r, column=1, value="TOP PRODUCTS PER SEASON (from PMCtire data)").font = SUB_FONT
    r += 1

    all_pmc = []
    for products in pmctire["sizes"].values():
        all_pmc.extend(products)

    season_groups = {"Winter": [], "All-Season / All-Weather": [], "Summer / Performance": []}
    for p in all_pmc:
        season = normalize_season(p.get("season"))
        if season == "Winter":
            season_groups["Winter"].append(p)
        elif "Summer" in season:
            season_groups["Summer / Performance"].append(p)
        else:
            season_groups["All-Season / All-Weather"].append(p)

    for group_name, products in season_groups.items():
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.cell(row=r, column=1, value=group_name).font = Font(bold=True, size=11)
        ws.cell(row=r, column=1).fill = SUB_FILL
        r += 1
        hdr_row(ws, r, ["Brand", "Model", "Size", "Price", "Margin %", "Category", "Vehicle Fitment", "Features"])
        r += 1
        seen = set()
        sorted_p = sorted(products, key=lambda x: x.get("price_cad") or 0, reverse=True)
        shown = 0
        for p in sorted_p:
            key = (normalize_brand(p["brand"]), p["model"])
            if key in seen:
                continue
            seen.add(key)
            margin = (p["price_cad"] - p["wholesale_cost"]) / p["price_cad"] if p.get("price_cad") and p.get("wholesale_cost") and p["price_cad"] > 0 else None
            ws.cell(row=r, column=1, value=normalize_brand(p["brand"])).font = Font(bold=True)
            ws.cell(row=r, column=2, value=p["model"])
            ws.cell(row=r, column=3, value=p["size"])
            ws.cell(row=r, column=4, value=p.get("price_cad"))
            ws.cell(row=r, column=4).number_format = CURR
            cell = ws.cell(row=r, column=5, value=margin)
            if margin is not None:
                cell.number_format = PCT
            ws.cell(row=r, column=6, value=p.get("category", ""))
            ws.cell(row=r, column=7, value=FITMENT.get(p["size"], ""))
            ws.cell(row=r, column=8, value=features_str(p))
            style_row(ws, r, 8)
            r += 1
            shown += 1
            if shown >= 12:
                break
        r += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 30
    for ci in range(3, 15):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    print("Seasonal Calendar: built")


def build_brand_rankings(wb, pmctire):
    ws = wb.create_sheet("Brand Rankings")
    ws.sheet_properties.tabColor = "BF8F00"
    ws.merge_cells("A1:K1")
    ws["A1"] = "BRAND RANKINGS \u2014 PMCtire Market Data"
    ws["A1"].font = TITLE_FONT
    total = sum(len(v) for v in pmctire["sizes"].values())
    ws.merge_cells("A2:K2")
    ws["A2"] = f"Based on {total:,} products across {len(pmctire['sizes'])} sizes | PMCtire.com | May 2026"
    ws["A2"].font = Font(italic=True, color="666666")

    brand_stats = defaultdict(lambda: {"prices": [], "costs": [], "sizes": set(), "models": set(), "categories": defaultdict(int), "seasons": defaultdict(int)})
    for size, products in pmctire["sizes"].items():
        for p in products:
            brand = normalize_brand(p["brand"])
            brand_stats[brand]["prices"].append(p.get("price_cad", 0))
            if p.get("wholesale_cost"):
                brand_stats[brand]["costs"].append(p["wholesale_cost"])
            brand_stats[brand]["sizes"].add(size)
            brand_stats[brand]["models"].add(p.get("model", ""))
            brand_stats[brand]["categories"][p.get("category", "")] += 1
            brand_stats[brand]["seasons"][normalize_season(p.get("season"))] += 1

    ranked = []
    for brand, stats in brand_stats.items():
        prices = [p for p in stats["prices"] if p]
        costs = stats["costs"]
        avg_price = sum(prices) / len(prices) if prices else 0
        avg_cost = sum(costs) / len(costs) if costs else 0
        avg_margin = (avg_price - avg_cost) / avg_price * 100 if avg_price > 0 and avg_cost > 0 else 0
        coverage = len(stats["sizes"]) / len(pmctire["sizes"]) * 100 if pmctire["sizes"] else 0
        score = (len(prices) / total * 100 * 0.4) + (coverage * 0.3) + (avg_margin * 0.3)
        ranked.append({
            "brand": brand, "score": round(score, 1), "listings": len(prices),
            "models": len(stats["models"]), "sizes": len(stats["sizes"]),
            "coverage": round(coverage, 1), "avg_price": round(avg_price, 2),
            "avg_cost": round(avg_cost, 2), "avg_margin": round(avg_margin, 1),
            "top_category": max(stats["categories"], key=stats["categories"].get) if stats["categories"] else "",
            "primary_season": max(stats["seasons"], key=stats["seasons"].get) if stats["seasons"] else "",
        })
    ranked.sort(key=lambda x: x["score"], reverse=True)

    headers = ["Rank", "Brand", "Score", "# Products", "# Models", "Size Coverage", "Avg Price", "Avg Cost", "Avg Margin %", "Top Category", "Primary Season"]
    r = 4
    hdr_row(ws, r, headers)

    notes = {
        "Michelin": "Premium leader. X-Ice Snow dominates winter. Made in Canada.",
        "Continental": "VikingContact 8 + TrueContact Tour. Strong OEM fitment.",
        "Bridgestone": "Blizzak winter leader. WeatherPeak strong all-weather.",
        "Goodyear": "Assurance line dominates all-season. WinterCommand for winter.",
        "Pirelli": "Premium performance focus. Scorpion line for SUV segment.",
        "Toyo": "Celsius II top all-weather. Strong value proposition.",
        "Hankook": "Budget-friendly quality. Winter i*pike RS2 popular studded.",
        "Yokohama": "Mid-market. Geolandar strong for SUV/truck.",
        "Cooper": "Strong mid-market value. Discoverer line for trucks.",
        "General": "Best value in premium. AltiMAX RT45 workhorse.",
        "Firestone": "Budget-friendly. Destination LE3 bestseller.",
        "Dunlop": "Winter Maxx 2 solid budget winter pick.",
        "BFGoodrich": "All-terrain specialist. Trail-Terrain T/A popular.",
        "Falken": "WildPeak A/T Trail strong crossover A/T.",
        "Nexen": "Value brand growing fast. Good size coverage.",
        "Nitto": "Performance/truck specialist. Motivo 365 all-weather.",
    }

    for i, b in enumerate(ranked[:20]):
        r += 1
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=b["brand"]).font = Font(bold=True, size=11)
        ws.cell(row=r, column=3, value=b["score"]).number_format = "0.0"
        ws.cell(row=r, column=4, value=b["listings"])
        ws.cell(row=r, column=5, value=b["models"])
        ws.cell(row=r, column=6, value=f"{b['sizes']}/{len(pmctire['sizes'])}")
        ws.cell(row=r, column=7, value=b["avg_price"]).number_format = CURR
        ws.cell(row=r, column=8, value=b["avg_cost"]).number_format = CURR
        ws.cell(row=r, column=9, value=b["avg_margin"] / 100).number_format = PCT
        ws.cell(row=r, column=10, value=b["top_category"])
        ws.cell(row=r, column=11, value=b["primary_season"])
        style_row(ws, r, 11)
        if i < 3:
            for ci in range(1, 12):
                ws.cell(row=r, column=ci).fill = MEDALS[i]

    r += 2
    categories = defaultdict(lambda: defaultdict(int))
    for products in pmctire["sizes"].values():
        for p in products:
            categories[p.get("category", "Other")][normalize_brand(p["brand"])] += 1

    for cat_name in ["Touring", "Performance", "All Terrain", "Highway Terrain", "Ultra Performance"]:
        if cat_name not in categories:
            continue
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(row=r, column=1, value=f"Top 5 \u2014 {cat_name}").font = SUB_FONT
        ws.cell(row=r, column=1).fill = SUB_FILL
        r += 1
        hdr_row(ws, r, ["Rank", "Brand", "Products", "Notes"])
        r += 1
        for j, (brand, count) in enumerate(sorted(categories[cat_name].items(), key=lambda x: -x[1])[:5]):
            ws.cell(row=r, column=1, value=j + 1)
            ws.cell(row=r, column=2, value=brand).font = Font(bold=True)
            ws.cell(row=r, column=3, value=count)
            ws.cell(row=r, column=4, value=notes.get(brand, ""))
            style_row(ws, r, 4)
            r += 1
        r += 1

    widths = [6, 16, 10, 10, 10, 12, 11, 11, 11, 16, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    print(f"Brand Rankings: {len(ranked)} brands")


def build_top_models(wb, pmctire):
    ws = wb.create_sheet("Top Models")
    ws.sheet_properties.tabColor = "7030A0"
    ws.merge_cells("A1:J1")
    ws["A1"] = "TOP MODELS \u2014 What People Actually Buy"
    ws["A1"].font = TITLE_FONT

    model_stats = defaultdict(lambda: {"sizes": set(), "prices": [], "costs": [], "categories": set(), "seasons": set(), "brand": ""})
    for size, products in pmctire["sizes"].items():
        for p in products:
            brand = normalize_brand(p["brand"])
            model = p.get("model", "")
            key = (brand, model)
            model_stats[key]["brand"] = brand
            model_stats[key]["sizes"].add(size)
            if p.get("price_cad"):
                model_stats[key]["prices"].append(p["price_cad"])
            if p.get("wholesale_cost"):
                model_stats[key]["costs"].append(p["wholesale_cost"])
            model_stats[key]["categories"].add(p.get("category", ""))
            model_stats[key]["seasons"].add(normalize_season(p.get("season")))

    models_list = []
    for (brand, model), stats in model_stats.items():
        prices = stats["prices"]
        costs = stats["costs"]
        avg_price = sum(prices) / len(prices) if prices else 0
        avg_cost = sum(costs) / len(costs) if costs else 0
        margin = (avg_price - avg_cost) / avg_price * 100 if avg_price > 0 and avg_cost > 0 else 0
        models_list.append({
            "brand": brand, "model": model, "sizes": len(stats["sizes"]),
            "avg_price": avg_price, "avg_cost": avg_cost, "margin": margin,
            "category": ", ".join(sorted(stats["categories"])),
            "season": ", ".join(sorted(stats["seasons"])),
        })
    models_list.sort(key=lambda x: x["sizes"], reverse=True)

    why_map = {
        ("Michelin", "X-Ice Snow"): "Canada's #1 winter tire. Made in Canada. Widest size range.",
        ("Michelin", "CrossClimate2"): "Best-reviewed all-weather. Year-round convenience.",
        ("Michelin", "Defender2"): "Top mileage warranty. Best daily driver tire.",
        ("Michelin", "Pilot Sport 4 S"): "Gold standard performance. OEM on luxury sports cars.",
        ("Continental", "VikingContact 8"): "Premium winter. EV optimized. Excellent ice grip.",
        ("Continental", "TrueContact Tour 54"): "80K warranty. Strong OEM fitment. Quiet ride.",
        ("Continental", "IceContact XTRM"): "Ultimate winter grip. Available studded or unstudded.",
        ("Bridgestone", "Blizzak WS90"): "Top winter all-rounder. Strong brand recognition.",
        ("Bridgestone", "WeatherPeak"): "Strong all-weather entry. 3PMSF + M+S rated.",
        ("Goodyear", "Assurance ComfortDrive"): "Quiet luxury ride. Strong in SUV segment.",
        ("Goodyear", "WinterCommand Ultra"): "Premium winter choice. Good ice performance.",
        ("Pirelli", "Scorpion All Season Plus 3"): "Premium SUV all-season. Pirelli quality.",
        ("Toyo", "Celsius II"): "Top all-weather pick. Great in unpredictable conditions.",
        ("Hankook", "Winter i*pike RS2 (W429)"): "Best value studded winter. Popular in Quebec.",
        ("Falken", "WildPeak A/T Trail"): "Top crossover A/T. 65K warranty.",
        ("General", "Altimax RT45"): "Budget king. 65K warranty, widest value coverage.",
        ("Nitto", "Motivo 365"): "Strong all-weather value. Good performance ratings.",
        ("BFGoodrich", "Trail-Terrain T/A"): "Popular crossover A/T. On-road comfort.",
    }

    headers = ["Rank", "Brand", "Model", "# Sizes", "Avg Price", "Avg Cost", "Margin %", "Category", "Season", "Why It Sells"]
    r = 3
    hdr_row(ws, r, headers)
    r += 1
    for i, m in enumerate(models_list[:60]):
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=m["brand"]).font = Font(bold=True)
        ws.cell(row=r, column=3, value=m["model"]).font = Font(bold=True)
        ws.cell(row=r, column=4, value=m["sizes"])
        ws.cell(row=r, column=5, value=m["avg_price"]).number_format = CURR
        ws.cell(row=r, column=6, value=m["avg_cost"]).number_format = CURR
        ws.cell(row=r, column=7, value=m["margin"] / 100).number_format = PCT
        ws.cell(row=r, column=8, value=m["category"])
        ws.cell(row=r, column=9, value=m["season"])
        ws.cell(row=r, column=10, value=why_map.get((m["brand"], m["model"]), ""))
        style_row(ws, r, 10)
        if i < 3:
            for ci in range(1, 11):
                ws.cell(row=r, column=ci).fill = MEDALS[i]
        r += 1

    widths = [6, 14, 30, 8, 10, 10, 10, 20, 20, 55]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "D4"
    print(f"Top Models: {min(60, len(models_list))} models")


def build_margin_analysis(wb, pmctire):
    ws = wb.create_sheet("Margin Analysis")
    ws.sheet_properties.tabColor = "C00000"
    ws.merge_cells("A1:H1")
    ws["A1"] = "MARGIN ANALYSIS \u2014 Wholesale Cost Intelligence"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:H2")
    ws["A2"] = "PMCtire wholesale costs reveal dealer margins | Use for competitive pricing strategy"
    ws["A2"].font = Font(italic=True, color="666666")

    all_products = []
    for size, products in pmctire["sizes"].items():
        for p in products:
            if p.get("price_cad") and p.get("wholesale_cost") and p["price_cad"] > 0:
                all_products.append({
                    "brand": normalize_brand(p["brand"]), "model": p["model"], "size": size,
                    "price": p["price_cad"], "cost": p["wholesale_cost"],
                    "margin_pct": (p["price_cad"] - p["wholesale_cost"]) / p["price_cad"] * 100,
                    "margin_dollar": p["price_cad"] - p["wholesale_cost"],
                    "category": p.get("category", ""),
                })

    r = 4
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="TOP 30 HIGHEST-MARGIN PRODUCTS").font = SUB_FONT
    ws.cell(row=r, column=1).fill = SUB_FILL
    r += 1
    hdr_row(ws, r, ["Brand", "Model", "Size", "Retail Price", "Wholesale Cost", "Margin $", "Margin %", "Category"])
    r += 1
    for p in sorted(all_products, key=lambda x: x["margin_pct"], reverse=True)[:30]:
        ws.cell(row=r, column=1, value=p["brand"]).font = Font(bold=True)
        ws.cell(row=r, column=2, value=p["model"])
        ws.cell(row=r, column=3, value=p["size"])
        ws.cell(row=r, column=4, value=p["price"]).number_format = CURR
        ws.cell(row=r, column=5, value=p["cost"]).number_format = CURR
        ws.cell(row=r, column=6, value=p["margin_dollar"]).number_format = CURR
        ws.cell(row=r, column=7, value=p["margin_pct"] / 100).number_format = PCT
        ws.cell(row=r, column=8, value=p["category"])
        style_row(ws, r, 8)
        r += 1

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value="AVERAGE MARGIN BY BRAND").font = SUB_FONT
    ws.cell(row=r, column=1).fill = SUB_FILL
    r += 1
    brand_margins = defaultdict(list)
    for p in all_products:
        brand_margins[p["brand"]].append(p["margin_pct"])
    hdr_row(ws, r, ["Brand", "Avg Margin %", "Min Margin", "Max Margin", "# Products", "Strategy"])
    r += 1
    strategies = [(0, 20, "Low margin \u2014 compete on volume"), (20, 28, "Moderate \u2014 solid everyday products"),
                  (28, 35, "Strong \u2014 push these products"), (35, 100, "Excellent \u2014 maximize sales")]
    for brand, margins in sorted(brand_margins.items(), key=lambda x: sum(x[1]) / len(x[1]), reverse=True):
        avg = sum(margins) / len(margins)
        strat = next((s for lo, hi, s in strategies if lo <= avg < hi), "")
        ws.cell(row=r, column=1, value=brand).font = Font(bold=True)
        ws.cell(row=r, column=2, value=avg / 100).number_format = PCT
        ws.cell(row=r, column=3, value=min(margins) / 100).number_format = PCT
        ws.cell(row=r, column=4, value=max(margins) / 100).number_format = PCT
        ws.cell(row=r, column=5, value=len(margins))
        ws.cell(row=r, column=6, value=strat)
        style_row(ws, r, 6)
        r += 1

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value="AVERAGE MARGIN BY SIZE (Top 20)").font = SUB_FONT
    ws.cell(row=r, column=1).fill = SUB_FILL
    r += 1
    size_margins = defaultdict(list)
    for p in all_products:
        size_margins[p["size"]].append(p["margin_pct"])
    hdr_row(ws, r, ["Size", "Vehicle Fitment", "Avg Margin %", "Avg Price", "# Products", "Opportunity"])
    r += 1
    for size, margins in sorted(size_margins.items(), key=lambda x: sum(x[1]) / len(x[1]), reverse=True)[:20]:
        avg_m = sum(margins) / len(margins)
        size_prods = [p for p in all_products if p["size"] == size]
        avg_price = sum(p["price"] for p in size_prods) / len(size_prods)
        opp = "HIGH" if avg_m > 30 and len(margins) >= 10 else ("GOOD" if avg_m > 25 else "MODERATE")
        ws.cell(row=r, column=1, value=size).font = Font(bold=True)
        ws.cell(row=r, column=2, value=FITMENT.get(size, ""))
        ws.cell(row=r, column=3, value=avg_m / 100).number_format = PCT
        ws.cell(row=r, column=4, value=avg_price).number_format = CURR
        ws.cell(row=r, column=5, value=len(margins))
        cell = ws.cell(row=r, column=6, value=opp)
        if opp == "HIGH":
            cell.font = Font(bold=True, color="006100")
        style_row(ws, r, 6)
        r += 1

    widths = [14, 28, 12, 12, 12, 10, 10, 50]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    print(f"Margin Analysis: {len(all_products)} products analyzed")


def build_size_opportunities(wb, pmctire):
    ws = wb.create_sheet("Size Opportunities")
    ws.sheet_properties.tabColor = "548235"
    ws.merge_cells("A1:J1")
    ws["A1"] = "SIZE OPPORTUNITIES \u2014 Where to Focus Inventory"
    ws["A1"].font = TITLE_FONT

    size_stats = []
    for size, products in sorted(pmctire["sizes"].items()):
        if not products:
            continue
        prices = [p["price_cad"] for p in products if p.get("price_cad")]
        costs = [p["wholesale_cost"] for p in products if p.get("wholesale_cost")]
        brands = set(normalize_brand(p["brand"]) for p in products)
        if not prices:
            continue
        avg_price = sum(prices) / len(prices)
        avg_cost = sum(costs) / len(costs) if costs else 0
        avg_margin = (avg_price - avg_cost) / avg_price * 100 if avg_price > 0 and avg_cost > 0 else 0
        opp_score = len(products) * avg_margin * (avg_price / 100)
        size_stats.append({
            "size": size, "products": len(products), "brands": len(brands),
            "min_price": min(prices), "avg_price": avg_price, "max_price": max(prices),
            "avg_margin": avg_margin, "opp_score": opp_score,
        })
    size_stats.sort(key=lambda x: x["opp_score"], reverse=True)

    r = 3
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.cell(row=r, column=1, value="ALL SIZES RANKED BY OPPORTUNITY SCORE").font = SUB_FONT
    ws.cell(row=r, column=1).fill = SUB_FILL
    r += 1
    headers = ["Rank", "Size", "Vehicle Fitment", "# Products", "# Brands", "Min Price", "Avg Price", "Max Price", "Avg Margin %", "Opportunity"]
    hdr_row(ws, r, headers)
    r += 1
    for i, s in enumerate(size_stats):
        opp = "\u2605\u2605\u2605" if i < 10 else ("\u2605\u2605" if i < 25 else "\u2605")
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=s["size"]).font = Font(bold=True)
        ws.cell(row=r, column=3, value=FITMENT.get(s["size"], ""))
        ws.cell(row=r, column=4, value=s["products"])
        ws.cell(row=r, column=5, value=s["brands"])
        ws.cell(row=r, column=6, value=s["min_price"]).number_format = CURR
        ws.cell(row=r, column=7, value=s["avg_price"]).number_format = CURR
        ws.cell(row=r, column=8, value=s["max_price"]).number_format = CURR
        ws.cell(row=r, column=9, value=s["avg_margin"] / 100).number_format = PCT
        ws.cell(row=r, column=10, value=opp)
        style_row(ws, r, 10)
        if i < 3:
            for ci in range(1, 11):
                ws.cell(row=r, column=ci).fill = MEDALS[i]
        r += 1

    widths = [6, 12, 42, 10, 10, 10, 10, 10, 10, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    print(f"Size Opportunities: {len(size_stats)} sizes ranked")


def main():
    print("Loading data...")
    gsheet, pmctire = load_data()
    if not pmctire["sizes"]:
        print("ERROR: No PMCtire data found. Run scrape_pmctire.py first.")
        sys.exit(1)
    print("\nBuilding Excel workbook...")
    wb = openpyxl.Workbook()
    build_master_catalogue(wb, gsheet, pmctire)
    build_price_intelligence(wb, pmctire, gsheet)
    build_seasonal_calendar(wb, pmctire)
    build_brand_rankings(wb, pmctire)
    build_top_models(wb, pmctire)
    build_margin_analysis(wb, pmctire)
    build_size_opportunities(wb, pmctire)
    out_path = os.path.join(MI_DIR, "Jun_Tire_Market_Intelligence_May2026.xlsx")
    wb.save(out_path)
    print(f"\nSaved: {out_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
