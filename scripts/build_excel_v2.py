import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from collections import defaultdict

rankings_path = "../market-intelligence/tire-brand-rankings.json"
retail_path = "../market-intelligence/retail-data.json"
output_path = "../market-intelligence/Canadian_Tire_Market_Rankings_May2026.xlsx"

with open(rankings_path) as f:
    rankings = json.load(f)
with open(retail_path) as f:
    retail = json.load(f)

# --- Flatten all listings ---
all_listings = []
for size, data in retail["sizes"].items():
    listings = data.get("listings", []) if isinstance(data, dict) else data
    for l in listings:
        all_listings.append(l)

# --- Parse size into components ---
def parse_size(size_str):
    m = re.match(r'(\d+)/(\d+)R(\d+)', size_str)
    if m:
        return int(m.group(1)), int(m.group(2)), int(m.group(3))
    return None, None, None

# --- Estimate performance tier from price ---
def estimate_tier(price):
    if price < 180:
        return "Value"
    elif price < 300:
        return "Mid"
    else:
        return "Premium"

# --- Map type to category ---
def normalize_category(t):
    t = (t or "unknown").lower()
    if "winter" in t:
        return "Winter"
    elif "all-season" in t or "all season" in t:
        return "All-Season"
    elif "summer" in t or "performance" in t:
        return "Summer"
    elif "terrain" in t:
        return "All-Terrain"
    return "All-Season"

# --- Season push logic ---
def season_push(cat):
    if cat == "Winter":
        return "Sep-Nov"
    elif cat == "Summer":
        return "Mar-May"
    return "Year-round (spring + fall peaks)"

# --- Vehicle fitment by size (common Canadian vehicles) ---
FITMENT = {
    "195/65R15": "Civic; Corolla; Sentra; Elantra; Mazda3",
    "205/55R16": "Civic; Corolla; Mazda3; Jetta; Elantra",
    "215/55R17": "Camry; Accord; Mazda6; Sonata; Altima",
    "215/60R16": "Camry; Accord; Sonata; Altima; Legacy",
    "215/65R16": "Escape; Tucson; Sportage; Forester; CX-5",
    "215/70R16": "CX-5; Forester; Outback; RAV4; CR-V",
    "225/40R18": "Civic Si; WRX; Golf GTI; Mazda3; Elantra N",
    "225/45R17": "Civic Si; Mazda3; WRX; Golf GTI; Sentra",
    "225/45R18": "Camry; Accord; Mazda6; Sonata; K5",
    "225/50R17": "Camry; Sonata; Altima; Malibu; Legacy",
    "225/55R17": "Camry; Accord; Mazda6; Sonata; Altima",
    "225/60R17": "RAV4; CR-V; Rogue; Escape; Forester",
    "225/60R18": "RAV4; CR-V; Tucson; Sportage; CX-5",
    "225/65R17": "RAV4; CR-V; Rogue; Escape; Outback",
    "235/45R18": "Camry; Accord; Mazda6; K5; Sonata",
    "235/55R18": "Highlander; Pilot; Palisade; Telluride; Explorer",
    "235/55R19": "Tucson; Sportage; CX-5; RAV4; CR-V",
    "235/60R18": "Highlander; Pilot; Palisade; Telluride; Explorer",
    "235/65R17": "Santa Fe; Tucson; Sportage; CX-5; Equinox",
    "235/65R18": "Santa Fe; Tucson; Sportage; CX-5; Equinox",
    "235/70R16": "Tacoma; Ranger; Canyon; Colorado; Frontier",
    "245/45R18": "Model 3; BMW 3; Audi A4; Mercedes C; IS",
    "245/45R19": "Model Y; BMW X3; Audi Q5; Mercedes GLC; XC60",
    "245/45R20": "Model Y; BMW X3; Audi Q5; Mercedes GLC; XC60",
    "245/60R18": "Traverse; Pilot; Explorer; Palisade; Sorento",
    "245/65R17": "4Runner; Wrangler; Bronco; Tacoma; Colorado",
    "245/70R16": "4Runner; Wrangler; Tacoma; Ranger; Colorado",
    "245/75R16": "Wrangler; 4Runner; Tacoma; Bronco; Gladiator",
    "255/35R19": "BMW M3; AMG C63; Audi S4; Corvette; Mustang GT",
    "255/45R19": "Model Y; BMW X3; Audi Q5; Mercedes GLC; XC60",
    "255/55R18": "Explorer; Grand Cherokee; Durango; Highlander; Pilot",
    "255/65R18": "F-150; RAM 1500; Silverado; Sierra; Tundra",
    "255/70R17": "F-150; RAM 1500; Silverado; Sierra; Tundra",
    "255/70R18": "F-150; RAM 1500; Silverado; Sierra; Tundra",
    "265/60R18": "F-150; RAM 1500; Silverado; Sierra; Frontier",
    "265/65R17": "F-150; RAM 1500; Silverado; Sierra; Tacoma",
    "265/65R18": "F-150; RAM 1500; Silverado; Sierra; Tundra",
    "265/70R17": "F-150; Silverado; Sierra; RAM 1500; Tacoma",
    "265/70R18": "F-150; RAM 1500; Silverado; Sierra; Tundra",
    "275/40R20": "BMW X5; Audi Q7; Mercedes GLE; Explorer ST; Grand Cherokee",
    "275/45R20": "F-150; RAM 1500; Tundra; Silverado; Sierra",
    "275/55R19": "F-150; RAM 1500; Tundra; Silverado; Sierra",
    "275/55R20": "F-150; RAM 1500; Tundra; Silverado; Sierra",
    "275/60R20": "F-150; RAM 1500; Tundra; Silverado; Sierra",
    "275/65R18": "F-150; RAM 1500; Silverado; Sierra; Tundra",
    "275/70R18": "F-150; RAM 1500; Silverado; Sierra; 2500HD",
    "285/45R22": "Escalade; Tahoe; Suburban; Yukon; Navigator",
    "285/65R18": "F-250; RAM 2500; Silverado 2500; Sierra 2500",
    "285/70R17": "F-250; RAM 2500; Silverado 2500; Sierra 2500; Super Duty",
    "305/35R24": "Escalade; Navigator; Suburban; Yukon XL",
    "175/65R15": "Fit; Yaris; Versa; Rio; Accent",
    "185/65R15": "Civic; Corolla; Sentra; Elantra; Forte",
    "215/45R17": "Civic; Corolla; Mazda3; Jetta; Elantra",
}

# --- Demand scores from rankings ---
demand_by_size = {}
for s in rankings["size_opportunities"]["highest_demand_sizes"]:
    demand_by_size[s["size"]] = s["demand_score"]

# --- Build per-size brand aggregation for pricing ---
size_brand_prices = defaultdict(lambda: defaultdict(list))
for l in all_listings:
    size_brand_prices[l["size"]][l["brand"]].append(l["price"])

# --- Priority tier logic ---
def priority_tier(demand, margin_pct):
    if demand >= 60 and margin_pct >= 30:
        return "T1"
    elif demand >= 30 or margin_pct >= 25:
        return "T2"
    return "T3"


wb = openpyxl.Workbook()

# --- Styling ---
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
data_font = Font(name="Calibri", size=10)
money_fmt = '$#,##0.00'
pct_fmt = '0.0%'
thin_border = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)

t1_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
t2_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
t3_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
tier_fills = {"T1": t1_fill, "T2": t2_fill, "T3": t3_fill}

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def style_data(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.border = thin_border

def auto_width(ws, max_w=28):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(mx + 3, max_w)

def freeze_and_filter(ws, header_row, last_col, last_row):
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = f"A{header_row + 1}"


# =====================================================
# SHEET 1: MASTER CATALOG (matches old spreadsheet)
# =====================================================
ws = wb.active
ws.title = "Master Catalog"

headers = [
    "Brand", "Model", "Size", "Width", "Aspect Ratio", "Rim Diameter",
    "Category", "Performance Tier",
    "Avg Retail Price", "Est. Wholesale (65%)", "Our Target Price",
    "Our Margin ($)", "Our Margin (%)",
    "Demand Score (1-100)", "Top Vehicle Fitment",
    "Priority Tier", "Season Push", "Retailer Found"
]
NCOLS = len(headers)

for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, NCOLS)

# Deduplicate by brand+model+size
seen = set()
row = 2
for l in sorted(all_listings, key=lambda x: (x["size"], x["brand"], x.get("model", ""))):
    key = (l["brand"], l.get("model", ""), l["size"])
    if key in seen:
        continue
    seen.add(key)

    width, aspect, rim = parse_size(l["size"])
    category = normalize_category(l.get("type", ""))
    tier = estimate_tier(l["price"])
    avg_retail = l["price"]
    wholesale = round(avg_retail * 0.65, 2)
    target = round(wholesale + 80, 2)
    margin_dollar = 80.00
    margin_pct = margin_dollar / target if target > 0 else 0
    demand = demand_by_size.get(l["size"], 25)
    fitment = FITMENT.get(l["size"], "")
    ptier = priority_tier(demand, margin_pct * 100)
    spush = season_push(category)

    vals = [
        l["brand"], l.get("model", ""), l["size"], width, aspect, rim,
        category, tier,
        avg_retail, wholesale, target,
        margin_dollar, margin_pct,
        demand, fitment,
        ptier, spush, l.get("retailer", "")
    ]
    for i, v in enumerate(vals, 1):
        ws.cell(row=row, column=i, value=v)
    style_data(ws, row, NCOLS)

    ws.cell(row=row, column=9).number_format = money_fmt
    ws.cell(row=row, column=10).number_format = money_fmt
    ws.cell(row=row, column=11).number_format = money_fmt
    ws.cell(row=row, column=12).number_format = money_fmt
    ws.cell(row=row, column=13).number_format = pct_fmt

    tier_cell = ws.cell(row=row, column=16)
    if ptier in tier_fills:
        tier_cell.fill = tier_fills[ptier]

    row += 1

last_row = row - 1
freeze_and_filter(ws, 1, NCOLS, last_row)
auto_width(ws)


# =====================================================
# SHEET 2: BRAND RANKINGS (Top 10 + full list)
# =====================================================
ws2 = wb.create_sheet("Brand Rankings")
headers2 = ["Rank", "Brand", "Composite Score", "Popularity Score", "Coverage Score",
            "Value Score", "Avg Price (CAD)", "Sizes Available (of 65)",
            "Performance Tier", "Key Strength"]

for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, len(headers2))

strengths = {
    "Continental": "Dominant in winter, widest winter range",
    "Toyo": "Strong all-around, great size coverage",
    "Michelin": "Widest coverage, highest reviews, premium positioning",
    "General": "Best value in top 5, strong winter presence",
    "Bridgestone": "Solid winter + all-season, good brand recognition",
    "Goodyear": "#1 all-season brand by listings",
    "Yokohama": "Good price point with wide coverage",
    "Firestone": "Budget-friendly premium, Bridgestone subsidiary",
    "Cooper": "Strong mid-market, good all-terrain options",
    "Pirelli": "Premium niche, strong performance segment",
}

for idx, b in enumerate(rankings["top_10_overall"], 1):
    r = idx + 1
    ws2.cell(row=r, column=1, value=idx)
    ws2.cell(row=r, column=2, value=b["brand"])
    ws2.cell(row=r, column=3, value=b["composite_score"])
    ws2.cell(row=r, column=4, value=b["popularity"])
    ws2.cell(row=r, column=5, value=b["coverage"])
    ws2.cell(row=r, column=6, value=b["value"])
    ws2.cell(row=r, column=7, value=b["avg_price"])
    ws2.cell(row=r, column=8, value=b["sizes_available"])
    ws2.cell(row=r, column=9, value=b["tier"].title())
    ws2.cell(row=r, column=10, value=strengths.get(b["brand"], ""))
    style_data(ws2, r, len(headers2))
    ws2.cell(row=r, column=7).number_format = money_fmt
    for c in [3, 4, 5, 6]:
        ws2.cell(row=r, column=c).number_format = '0.0'
    if idx <= 3:
        for c in range(1, len(headers2) + 1):
            ws2.cell(row=r, column=c).font = Font(name="Calibri", size=10, bold=True)

freeze_and_filter(ws2, 1, len(headers2), 11)
auto_width(ws2)


# =====================================================
# SHEET 3: PRICE INTELLIGENCE BY SIZE
# =====================================================
ws3 = wb.create_sheet("Price Intelligence by Size")
headers3 = ["Size", "Width", "Aspect", "Rim", "# Brands", "# Listings",
            "Min Price", "Max Price", "Avg Price", "Price Spread",
            "Demand Score", "Top Vehicle Fitment",
            "Top Brand 1", "Top Brand 2", "Top Brand 3"]

for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, len(headers3))

size_stats = {}
for size, data in retail["sizes"].items():
    listings = data.get("listings", []) if isinstance(data, dict) else data
    if not listings:
        continue
    prices = [l["price"] for l in listings if l.get("price")]
    brands = list(set(l["brand"] for l in listings))
    brand_counts = defaultdict(int)
    for l in listings:
        brand_counts[l["brand"]] += 1
    top_brands = sorted(brand_counts, key=brand_counts.get, reverse=True)

    size_stats[size] = {
        "listings": len(listings),
        "brands": len(brands),
        "min_p": min(prices) if prices else 0,
        "max_p": max(prices) if prices else 0,
        "avg_p": sum(prices) / len(prices) if prices else 0,
        "top_brands": top_brands[:3],
    }

row = 2
for size in sorted(size_stats.keys(), key=lambda s: (parse_size(s)[2] or 0, parse_size(s)[0] or 0)):
    s = size_stats[size]
    w, a, r_d = parse_size(size)
    demand = demand_by_size.get(size, "")

    vals = [size, w, a, r_d, s["brands"], s["listings"],
            s["min_p"], s["max_p"], round(s["avg_p"], 2), round(s["max_p"] - s["min_p"], 2),
            demand, FITMENT.get(size, "")]
    vals += s["top_brands"] + [""] * (3 - len(s["top_brands"]))

    for i, v in enumerate(vals, 1):
        ws3.cell(row=row, column=i, value=v)
    style_data(ws3, row, len(headers3))
    for c in [7, 8, 9, 10]:
        ws3.cell(row=row, column=c).number_format = money_fmt
    row += 1

freeze_and_filter(ws3, 1, len(headers3), row - 1)
auto_width(ws3)


# =====================================================
# SHEET 4: BY CATEGORY
# =====================================================
ws4 = wb.create_sheet("Category Breakdown")
headers4 = ["Category", "Rank", "Brand", "Listings", "Market Share"]

for i, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, len(headers4))

cat_fills = {
    "All-Season": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "Winter": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "Performance": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
}

row = 2
for cat_key, cat_label in [("all_season", "All-Season"), ("winter", "Winter"), ("performance", "Performance")]:
    brands = rankings["by_category"].get(cat_key, [])
    total = sum(b["listings"] for b in brands) if brands else 1
    for idx, b in enumerate(brands, 1):
        ws4.cell(row=row, column=1, value=cat_label)
        ws4.cell(row=row, column=2, value=idx)
        ws4.cell(row=row, column=3, value=b["brand"])
        ws4.cell(row=row, column=4, value=b["listings"])
        ws4.cell(row=row, column=5, value=round(b["listings"] / total, 3))
        ws4.cell(row=row, column=5).number_format = pct_fmt
        style_data(ws4, row, len(headers4))
        if cat_label in cat_fills:
            ws4.cell(row=row, column=1).fill = cat_fills[cat_label]
        row += 1

freeze_and_filter(ws4, 1, len(headers4), row - 1)
auto_width(ws4)


# =====================================================
# SHEET 5: BEST MARGIN OPPORTUNITIES
# =====================================================
ws5 = wb.create_sheet("Margin Opportunities")
headers5 = ["Rank", "Size", "Avg Price (CAD)", "Price Range", "Price Spread",
            "# Brands", "Demand Score", "Vehicle Fitment", "Opportunity Notes"]

for i, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, len(headers5))

for idx, s in enumerate(rankings["size_opportunities"]["best_margin_sizes"], 1):
    r = idx + 1
    demand = demand_by_size.get(s["size"], "")
    note = ""
    if s["brand_count"] >= 15:
        note = "High competition — many brands available"
    elif s["avg_price"] > 400:
        note = "High-value segment — premium pricing"

    ws5.cell(row=r, column=1, value=idx)
    ws5.cell(row=r, column=2, value=s["size"])
    ws5.cell(row=r, column=3, value=s["avg_price"])
    ws5.cell(row=r, column=4, value=s["price_range"])
    ws5.cell(row=r, column=5, value=s["margin_score"])
    ws5.cell(row=r, column=6, value=s["brand_count"])
    ws5.cell(row=r, column=7, value=demand)
    ws5.cell(row=r, column=8, value=FITMENT.get(s["size"], ""))
    ws5.cell(row=r, column=9, value=note)
    style_data(ws5, r, len(headers5))
    ws5.cell(row=r, column=3).number_format = money_fmt
    ws5.cell(row=r, column=5).number_format = '#,##0.0'

freeze_and_filter(ws5, 1, len(headers5), 11)
auto_width(ws5)


# =====================================================
# SHEET 6: SEASONAL CALENDAR
# =====================================================
ws6 = wb.create_sheet("Seasonal Calendar")
headers6 = ["Season", "Months", "Push Categories", "Key Actions", "Top Brands to Feature"]

for i, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=i, value=h)
style_header(ws6, 1, len(headers6))

calendar = [
    ["Spring", "Mar-May", "All-Season, Summer, All-Terrain",
     "Push all-season changeovers; promote summer performance; market all-terrain for camping season",
     "Michelin, Continental, Goodyear, Toyo, Bridgestone"],
    ["Summer", "Jun-Aug", "All-Season, All-Terrain",
     "Maintain all-season stock; push all-terrain for road trips; early winter pre-orders",
     "Continental, Toyo, General, Yokohama, Cooper"],
    ["Fall", "Sep-Nov", "Winter, All-Season",
     "PEAK SEASON: Winter tire changeover rush; push winter packages; all-season for holdouts",
     "Continental, Toyo, General, Bridgestone, Michelin"],
    ["Winter", "Dec-Feb", "Winter",
     "Late winter buyers; maintain winter stock; start planning spring inventory",
     "Continental, Toyo, General, Bridgestone, Goodyear"],
]

for idx, cal in enumerate(calendar):
    r = idx + 2
    for i, v in enumerate(cal, 1):
        ws6.cell(row=r, column=i, value=v)
    style_data(ws6, r, len(headers6))
    ws6.cell(row=r, column=1).font = Font(name="Calibri", size=10, bold=True)

auto_width(ws6, max_w=50)


# =====================================================
# SHEET 7: BUDGET ALTERNATIVES
# =====================================================
ws7 = wb.create_sheet("Budget Alternatives")
headers7 = ["Brand", "Avg Price (CAD)", "Value Score", "Price Consistency",
            "Tier", "Sizes Available", "Best Use Case"]

for i, h in enumerate(headers7, 1):
    ws7.cell(row=1, column=i, value=h)
style_header(ws7, 1, len(headers7))

budget_notes = {
    "Ironman": "Fleet / high-volume value buyers",
    "Westlake": "Budget-conscious consumers, wide size range",
}

row = 2
for tier_key in ["midrange", "budget"]:
    for b in rankings["by_value_tier"].get(tier_key, []):
        cov = next((c for c in rankings["by_coverage"] if c["brand"] == b["brand"]), None)
        sizes = cov["sizes"] if cov else ""
        ws7.cell(row=row, column=1, value=b["brand"])
        ws7.cell(row=row, column=2, value=b["avg_price"])
        ws7.cell(row=row, column=3, value=b["score"])
        ws7.cell(row=row, column=4, value=f"{b['consistency']:.0%}")
        ws7.cell(row=row, column=5, value=tier_key.title())
        ws7.cell(row=row, column=6, value=sizes)
        ws7.cell(row=row, column=7, value=budget_notes.get(b["brand"], ""))
        style_data(ws7, row, len(headers7))
        ws7.cell(row=row, column=2).number_format = money_fmt
        ws7.cell(row=row, column=3).number_format = '0.0'
        row += 1

auto_width(ws7)


# =====================================================
# SHEET 8: HIGHEST DEMAND SIZES
# =====================================================
ws8 = wb.create_sheet("Highest Demand Sizes")
headers8 = ["Rank", "Size", "Demand Score", "Avg Price (CAD)", "# Brands",
            "Vehicle Fitment", "Recommended Brands"]

for i, h in enumerate(headers8, 1):
    ws8.cell(row=1, column=i, value=h)
style_header(ws8, 1, len(headers8))

for idx, s in enumerate(rankings["size_opportunities"]["highest_demand_sizes"], 1):
    r = idx + 1
    # Get top 3 brands for this size
    sb = size_brand_prices.get(s["size"], {})
    top3 = sorted(sb.keys(), key=lambda b: len(sb[b]), reverse=True)[:3]

    ws8.cell(row=r, column=1, value=idx)
    ws8.cell(row=r, column=2, value=s["size"])
    ws8.cell(row=r, column=3, value=s["demand_score"])
    ws8.cell(row=r, column=4, value=s["avg_price"])
    ws8.cell(row=r, column=5, value=s["brand_count"])
    ws8.cell(row=r, column=6, value=FITMENT.get(s["size"], ""))
    ws8.cell(row=r, column=7, value="; ".join(top3))
    style_data(ws8, r, len(headers8))
    ws8.cell(row=r, column=4).number_format = money_fmt

freeze_and_filter(ws8, 1, len(headers8), 11)
auto_width(ws8)


# =====================================================
# SHEET 9: RAW LISTINGS DATA
# =====================================================
ws9 = wb.create_sheet("All Listings (Raw)")
headers9 = ["Size", "Brand", "Model", "Price (CAD)", "Category", "Retailer",
            "Width", "Aspect Ratio", "Rim Diameter", "Performance Tier"]

for i, h in enumerate(headers9, 1):
    ws9.cell(row=1, column=i, value=h)
style_header(ws9, 1, len(headers9))

row = 2
for l in sorted(all_listings, key=lambda x: (x["size"], x["brand"], x.get("price", 0))):
    w, a, rd = parse_size(l["size"])
    cat = normalize_category(l.get("type", ""))
    tier = estimate_tier(l.get("price", 0))

    ws9.cell(row=row, column=1, value=l["size"])
    ws9.cell(row=row, column=2, value=l["brand"])
    ws9.cell(row=row, column=3, value=l.get("model", ""))
    ws9.cell(row=row, column=4, value=l.get("price", 0))
    ws9.cell(row=row, column=5, value=cat)
    ws9.cell(row=row, column=6, value=l.get("retailer", "unknown"))
    ws9.cell(row=row, column=7, value=w)
    ws9.cell(row=row, column=8, value=a)
    ws9.cell(row=row, column=9, value=rd)
    ws9.cell(row=row, column=10, value=tier)
    style_data(ws9, row, len(headers9))
    ws9.cell(row=row, column=4).number_format = money_fmt
    row += 1

freeze_and_filter(ws9, 1, len(headers9), row - 1)
auto_width(ws9)


wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Master Catalog: {last_row} rows")
print(f"Raw Listings: {row - 2} rows")
print(f"Price Intelligence: {len(size_stats)} sizes")
