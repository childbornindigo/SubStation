#!/usr/bin/env python3
"""
build_final_excel.py — Canadian Tire Market Intelligence Spreadsheet Builder

Reads:
  - market-intelligence/tire-specs.json   (tiresize.com spec data, optional)
  - market-intelligence/retail-data.json  (Bing Shopping price data)

Produces:
  - market-intelligence/Canadian_Tire_Market_Intelligence_May2026.xlsx

Tabs:
  1. Master Catalog        — every brand/model/size with specs, prices, margins
  2. Price Intelligence     — per-size competitive landscape
  3. Seasonal Calendar      — 12-month push plan
  4. Brand Summary          — per-brand overview
  5. Top Models by Demand   — top 50 ranked by demand score
  6. All Listings (Raw)     — every Bing listing verbatim
"""

import json
import os
import re
import statistics
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE = Path(__file__).resolve().parent.parent / "market-intelligence"
SPECS_PATH = BASE / "tire-specs.json"
RETAIL_PATH = BASE / "retail-data.json"
OUTPUT_PATH = BASE / "Canadian_Tire_Market_Intelligence_May2026.xlsx"

# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------
specs_data = []
if SPECS_PATH.exists():
    with open(SPECS_PATH) as f:
        raw_specs = json.load(f)
        # Accept either a list or a dict with a "tires" key
        if isinstance(raw_specs, list):
            specs_data = raw_specs
        elif isinstance(raw_specs, dict):
            specs_data = raw_specs.get("tires", raw_specs.get("data", []))
            if isinstance(specs_data, dict):
                # Flatten if keyed by size
                flat = []
                for size_key, entries in specs_data.items():
                    if isinstance(entries, list):
                        flat.extend(entries)
                specs_data = flat
    print(f"Loaded {len(specs_data)} spec entries from tire-specs.json")
else:
    print("tire-specs.json not found — proceeding with price data only")

with open(RETAIL_PATH) as f:
    retail = json.load(f)

# ---------------------------------------------------------------------------
# Vehicle fitment map (~55 common Canadian sizes)
# ---------------------------------------------------------------------------
FITMENT = {
    "175/65R15": "Fit, Yaris, Versa, Rio, Accent",
    "175/70R14": "Fit, Yaris, Versa, Spark",
    "185/55R16": "Yaris, Fit, Fiesta, Accent",
    "185/60R15": "Civic, Corolla, Sentra, Elantra",
    "185/65R15": "Civic, Corolla, Sentra, Elantra, Forte",
    "195/50R16": "Prius, Corolla Hybrid, Insight",
    "195/55R16": "Civic, Corolla, Jetta, Elantra",
    "195/60R15": "Corolla, Civic, Sentra, Mazda3",
    "195/65R15": "Civic, Corolla, Sentra, Elantra, Mazda3",
    "205/45R17": "Civic, Corolla, Mazda3, Golf",
    "205/50R17": "Jetta, Civic, Corolla, Mazda3",
    "205/55R16": "Civic, Corolla, Mazda3, Jetta, Elantra",
    "205/60R16": "Camry, Accord, Sonata, Altima, Legacy",
    "205/65R16": "Escape, Rogue, Sportage, Tucson, Forester",
    "215/45R17": "Corolla, Mazda3, Civic, Elantra",
    "215/50R17": "Camry, Accord, Mazda6, Altima",
    "215/55R17": "Camry, Accord, Mazda6, Sonata, Altima",
    "215/60R16": "Camry, Accord, Sonata, Altima, Legacy",
    "215/65R16": "Escape, Tucson, Sportage, Forester, CX-5",
    "215/65R17": "CR-V, Escape, Rogue, Tucson, Forester",
    "215/70R16": "CX-5, Forester, Outback, RAV4, CR-V",
    "225/40R18": "Civic Si, WRX, Golf GTI, Mazda3, Elantra N",
    "225/45R17": "Civic Si, WRX, GTI, Mazda3, Sentra",
    "225/45R18": "Camry, Accord, Mazda6, Sonata, K5",
    "225/50R17": "Camry, Sonata, Altima, Malibu, Legacy",
    "225/50R18": "Edge, Outback, Venza, CX-5",
    "225/55R17": "Accord, Camry, Maxima, Optima, Mazda6",
    "225/55R18": "Edge, Outback, Venza, CX-5, Murano",
    "225/60R17": "RAV4, CR-V, Rogue, Escape, Forester",
    "225/60R18": "RAV4, CR-V, Tucson, Sportage, CX-5",
    "225/65R17": "RAV4, CR-V, Rogue, Escape, Outback",
    "235/40R19": "Model 3, BMW 3, Audi A4, C-Class",
    "235/45R18": "Tesla 3, BMW 3, Lexus IS, WRX",
    "235/50R18": "Tesla Y, Model 3, Polestar, Volvo S60",
    "235/55R18": "Highlander, Pilot, Palisade, Telluride, Explorer",
    "235/55R19": "Tucson, Sportage, CX-5, RAV4, CR-V",
    "235/60R18": "Highlander, Pilot, Explorer, Santa Fe, Palisade",
    "235/65R17": "Santa Fe, Tucson, Sportage, CX-5, Equinox",
    "235/65R18": "Santa Fe, Tucson, Sportage, CX-5, Equinox",
    "235/70R16": "Tacoma, Ranger, Canyon, Colorado, Frontier",
    "245/35R20": "BMW M3, AMG C63, Audi S5, Corvette",
    "245/40R18": "BMW 3, IS350, WRX STI, Golf R",
    "245/45R18": "Model 3, BMW 3, Audi A4, Mercedes C, IS",
    "245/55R19": "Model Y, BMW X3, Audi Q5, XC60, GLC",
    "245/60R18": "Traverse, Explorer, Grand Cherokee, Durango, Sorento",
    "245/65R17": "4Runner, Wrangler, Bronco, Tacoma, Colorado",
    "245/75R16": "Wrangler, 4Runner, Tacoma, Bronco, Gladiator",
    "255/35R19": "BMW M3, AMG C63, Audi S4, Corvette, Mustang GT",
    "255/45R19": "Model Y, BMW X3, Audi Q5, Mercedes GLC, XC60",
    "255/50R19": "Cayenne, X5, GLE Coupe, XC90",
    "255/55R18": "Explorer, Grand Cherokee, Durango, Highlander, Pilot",
    "255/65R18": "F-150, RAM 1500, Silverado, Sierra, Tundra",
    "255/70R18": "F-150, RAM 1500, Silverado, Sierra, Tundra",
    "265/50R20": "Tahoe, Suburban, Yukon, Sequoia, Expedition",
    "265/60R18": "F-150, RAM 1500, Silverado, Sierra, Tundra",
    "265/60R20": "Tahoe, Yukon, Expedition, Sequoia, Armada",
    "265/65R18": "F-150, RAM 1500, Silverado, Sierra, Tundra",
    "265/70R16": "Tacoma, Ranger, Colorado, Frontier, Canyon",
    "265/70R17": "F-150, Silverado, Sierra, RAM 1500, Tacoma",
    "265/75R16": "F-250, RAM 2500, Silverado 2500, Sierra 2500",
    "275/40R20": "BMW X5, Audi Q7, Mercedes GLE, Explorer ST, Grand Cherokee",
    "275/45R20": "F-150, RAM 1500, Tundra, Silverado, Sierra",
    "275/55R19": "F-150, RAM 1500, Explorer, Palisade, Telluride",
    "275/55R20": "F-150, RAM 1500, Tundra, Silverado, Sierra",
    "275/60R20": "F-150, RAM 1500, Tundra, Expedition, Yukon",
    "275/65R18": "F-150, RAM 1500, Silverado, Sierra, Tundra",
    "275/70R18": "F-250, RAM 2500, Silverado 2500, Sierra 2500",
    "285/45R22": "Escalade, Tahoe, Suburban, Yukon, Navigator",
    "285/65R18": "F-250, RAM 2500, Silverado 2500, Sierra 2500",
    "285/70R17": "F-250, RAM 2500, Silverado 2500, Super Duty",
    "285/75R16": "F-250, RAM 2500, Silverado 2500, Super Duty",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_size(size_str):
    """Parse '225/65R17' -> (225, 65, 17). Returns (None, None, None) on fail."""
    m = re.match(r"(\d{3})/(\d{2,3})R(\d{2})", str(size_str))
    if m:
        return int(m.group(1)), int(m.group(2)), int(m.group(3))
    return None, None, None


def clean_model_name(brand, model):
    """Remove Bing Shopping junk from model strings."""
    if not model:
        return ""
    m = model.strip()

    # Remove trailing unicode ellipsis
    m = re.sub(r"[\u2026…]+$", "", m).strip()

    # Remove embedded size strings like "195/65R15" or "195/65 R15" or "195-65R15"
    m = re.sub(r"\d{3}[/ -]\d{2,3}\s*R?\d{2,3}", "", m).strip()

    # Remove "Tire By <brand>" junk (with or without text after "By")
    m = re.sub(r"\s*Tire\s+By.*$", "", m, flags=re.IGNORECASE).strip()

    # Remove trailing "Radial Tire" / "Tire"
    m = re.sub(r"\s+Radial\s+Tire\s*$", "", m, flags=re.IGNORECASE).strip()
    m = re.sub(r"\s+Tire\s*$", "", m, flags=re.IGNORECASE).strip()

    # Remove generic "ALL SEASON" / "ALL-SEASON" / "WINTER" only model names
    if re.match(r"^(ALL[- ]?SEASON|WINTER|SUMMER|ALL[- ]?TERRAIN)\s*$", m, re.IGNORECASE):
        return ""

    # Remove brand prefix if model starts with it
    if brand and m.lower().startswith(brand.lower()):
        m = m[len(brand):].strip()

    # Remove leading dashes/spaces
    m = re.sub(r"^[-–\s]+", "", m).strip()

    # Collapse whitespace
    m = re.sub(r"\s+", " ", m)

    # Remove trailing dash/junk
    m = re.sub(r"[-–\s]+$", "", m).strip()

    return m if len(m) >= 2 else ""


def classify_category(tire_type, model_name=""):
    """Classify into: Winter, All-Season, Performance, All-Terrain, Summer."""
    model_lower = (model_name or "").lower()
    t = (tire_type or "").lower()

    winter_kw = [
        "winter", "blizzak", "ice", "snow", "winguard", "wintercraft",
        "x-ice", "observe", "alpin", "hakkapeliitta", "ice zero",
        "ice friction", "wintercommand", "winterforce", "ice blazer",
        "studda", "sw606", "sw608",
    ]
    at_kw = [
        "all-terrain", "all terrain", "a/t", "grabber", "wildpeak",
        "discoverer", "open country a/t", "trail", "dynapro",
        "geolandar", "terra grappler",
    ]
    perf_kw = [
        "sport", "pilot", "potenza", "p zero", "ventus", "advan",
        "eagle f1", "contisport", "ps4", "ps5", "proxes sport",
        "pzero", "s-04",
    ]
    summer_kw = [
        "summer", "proxes r1", "re71", "ad09", "rt660", "re-71",
    ]

    if t == "winter" or any(k in model_lower for k in winter_kw):
        return "Winter"
    if any(k in model_lower for k in at_kw):
        return "All-Terrain"
    if t == "performance" or any(k in model_lower for k in perf_kw):
        return "Performance"
    if any(k in model_lower for k in summer_kw):
        return "Summer"
    if t == "all-season" or "all season" in t or "all-season" in model_lower:
        return "All-Season"
    return "All-Season"  # default


def performance_tier(avg_price):
    """Value / Mid / Premium based on price."""
    if avg_price is None:
        return ""
    if avg_price >= 300:
        return "Premium"
    if avg_price >= 180:
        return "Mid"
    return "Value"


def season_push(category):
    if category == "Winter":
        return "Sep-Nov"
    if category in ("Performance", "Summer"):
        return "Mar-May"
    return "Year-round"


def fuzzy_model_match(spec_model, listing_model):
    """
    Brand must already be exact-matched by caller.
    Fuzzy match: check if the cleaned listing model is a prefix/substring of
    the spec model or vice versa (Bing truncates names).
    """
    if not spec_model or not listing_model:
        return False

    s = spec_model.lower().strip()
    l = listing_model.lower().strip()

    # Exact
    if s == l:
        return True

    # One starts with the other (handles truncation)
    if s.startswith(l) or l.startswith(s):
        return True

    # First word match (e.g. "Endeavor" vs "Endeavor All-Season")
    s_first = s.split()[0] if s else ""
    l_first = l.split()[0] if l else ""
    if len(s_first) >= 3 and s_first == l_first:
        # Check second word too if available
        s_words = s.split()
        l_words = l.split()
        if len(s_words) >= 2 and len(l_words) >= 2:
            return s_words[1] == l_words[1]
        return True

    # Substring for short model names (like "GR906")
    if len(s) >= 4 and s in l:
        return True
    if len(l) >= 4 and l in s:
        return True

    return False


# ---------------------------------------------------------------------------
# Flatten all retail listings
# ---------------------------------------------------------------------------
all_listings = []
for size, data in retail["sizes"].items():
    listings = data.get("listings", []) if isinstance(data, dict) else data
    for l in listings:
        all_listings.append(l)

print(f"Loaded {len(all_listings)} retail listings across {len(retail['sizes'])} sizes")

# ---------------------------------------------------------------------------
# Build spec lookup: {(brand_lower, model_lower): spec_dict}
# ---------------------------------------------------------------------------
spec_lookup = {}
for spec in specs_data:
    brand = (spec.get("brand") or "").strip()
    model = (spec.get("model") or "").strip()
    if brand and model:
        key = (brand.lower(), model.lower())
        spec_lookup[key] = spec

print(f"Spec lookup entries: {len(spec_lookup)}")

# ---------------------------------------------------------------------------
# Group listings by (brand, cleaned_model, size)
# ---------------------------------------------------------------------------
listing_groups = defaultdict(list)
for l in all_listings:
    brand = l["brand"]
    raw_model = l.get("model", "")
    cleaned = clean_model_name(brand, raw_model)
    size = l.get("size", "")
    # Use cleaned name, or fall back to "Unspecified" for junk models
    model_key = cleaned if cleaned else "Unspecified"
    listing_groups[(brand, model_key, size)].append(l)

# ---------------------------------------------------------------------------
# Cross-reference: merge specs + prices into master catalog entries
# ---------------------------------------------------------------------------
catalog = []
matched_spec_keys = set()

for (brand, model, size), listings in listing_groups.items():
    prices = [l["price"] for l in listings if l.get("price") and l["price"] > 0]
    avg_price = round(statistics.mean(prices), 2) if prices else None
    min_price = min(prices) if prices else None
    max_price = max(prices) if prices else None

    width, aspect, rim = parse_size(size)

    # Try to find matching spec
    matched_spec = None
    brand_lower = brand.lower()

    # Exact match first
    exact_key = (brand_lower, model.lower())
    if exact_key in spec_lookup:
        matched_spec = spec_lookup[exact_key]
        matched_spec_keys.add(exact_key)
    else:
        # Fuzzy match: same brand, partial model
        for sk, sv in spec_lookup.items():
            if sk[0] == brand_lower and fuzzy_model_match(sk[1], model):
                matched_spec = sv
                matched_spec_keys.add(sk)
                break

    # Determine category from multiple sources
    listing_type = listings[0].get("type", "unknown")
    spec_category = matched_spec.get("category", "") if matched_spec else ""
    category = spec_category if spec_category else classify_category(listing_type, model)

    # Spec fields
    speed_rating = matched_spec.get("speed_rating", "") if matched_spec else ""
    load_index = matched_spec.get("load_index", "") if matched_spec else ""
    warranty_miles = matched_spec.get("warranty_miles", None) if matched_spec else None

    tier = performance_tier(avg_price)
    wholesale = round(avg_price * 0.65, 2) if avg_price else None
    target = round(avg_price * 0.85, 2) if avg_price else None
    margin_dollar = round(target - wholesale, 2) if (target and wholesale) else None
    margin_pct = round(margin_dollar / target, 4) if (target and margin_dollar and target > 0) else None

    fitment = FITMENT.get(size, "")
    listing_count = len(listings)

    catalog.append({
        "brand": brand,
        "model": model,
        "size": size,
        "width": width,
        "aspect_ratio": aspect,
        "rim_diameter": rim,
        "category": category,
        "performance_tier": tier,
        "speed_rating": speed_rating,
        "load_index": load_index,
        "warranty_miles": warranty_miles,
        "avg_price": avg_price,
        "min_price": min_price,
        "max_price": max_price,
        "wholesale": wholesale,
        "target_price": target,
        "margin_dollar": margin_dollar,
        "margin_pct": margin_pct,
        "fitment": fitment,
        "listing_count": listing_count,
        "season_push": season_push(category),
        "has_specs": matched_spec is not None,
        "has_prices": avg_price is not None,
    })

# Add spec-only entries (not matched to any price listing)
for sk, sv in spec_lookup.items():
    if sk in matched_spec_keys:
        continue
    brand = sv.get("brand", "")
    model = sv.get("model", "")
    size = sv.get("size", "")
    width, aspect, rim = parse_size(size)
    category = sv.get("category", "All-Season")
    catalog.append({
        "brand": brand,
        "model": model,
        "size": size,
        "width": width,
        "aspect_ratio": aspect,
        "rim_diameter": rim,
        "category": category,
        "performance_tier": "",
        "speed_rating": sv.get("speed_rating", ""),
        "load_index": sv.get("load_index", ""),
        "warranty_miles": sv.get("warranty_miles", None),
        "avg_price": None,
        "min_price": None,
        "max_price": None,
        "wholesale": None,
        "target_price": None,
        "margin_dollar": None,
        "margin_pct": None,
        "fitment": FITMENT.get(size, ""),
        "listing_count": 0,
        "season_push": season_push(category),
        "has_specs": True,
        "has_prices": False,
    })

# ---------------------------------------------------------------------------
# Demand score: composite of listing frequency + size popularity + brand breadth
# ---------------------------------------------------------------------------
# Factor 1: listing count for this brand+model across all sizes
brand_model_counts = defaultdict(int)
for c in catalog:
    brand_model_counts[(c["brand"], c["model"])] += c["listing_count"]
max_bm_count = max(brand_model_counts.values()) if brand_model_counts else 1

# Factor 2: how popular is this size? (total listings in that size)
size_listing_counts = defaultdict(int)
for c in catalog:
    size_listing_counts[c["size"]] += c["listing_count"]
max_size_count = max(size_listing_counts.values()) if size_listing_counts else 1

# Factor 3: number of sizes this brand+model spans
brand_model_sizes = defaultdict(set)
for c in catalog:
    if c["listing_count"] > 0:
        brand_model_sizes[(c["brand"], c["model"])].add(c["size"])

for c in catalog:
    bm_key = (c["brand"], c["model"])
    # Listing frequency (40% weight)
    freq_score = (brand_model_counts[bm_key] / max_bm_count) * 100
    # Size popularity (30% weight)
    size_score = (size_listing_counts.get(c["size"], 0) / max_size_count) * 100
    # Size breadth (30% weight) — models available in many sizes are more in-demand
    breadth = len(brand_model_sizes.get(bm_key, set()))
    breadth_score = min(100, breadth * 10)  # 10 sizes = 100

    composite = freq_score * 0.40 + size_score * 0.30 + breadth_score * 0.30
    c["demand_score"] = max(1, round(composite)) if c["listing_count"] > 0 else 0

# ---------------------------------------------------------------------------
# Priority tier: T1 = high demand + good margin, T2 = moderate, T3 = low
# Uses percentile-based thresholds for meaningful distribution
# ---------------------------------------------------------------------------
scored = sorted([c["demand_score"] for c in catalog if c["demand_score"] > 0])
p66 = scored[int(len(scored) * 0.66)] if scored else 50
p33 = scored[int(len(scored) * 0.33)] if scored else 25

for c in catalog:
    d = c["demand_score"]
    has_price = c["avg_price"] is not None
    if d >= p66 and has_price:
        c["priority"] = "T1"
    elif d >= p33:
        c["priority"] = "T2"
    else:
        c["priority"] = "T3"

# Sort: T1 first, then by demand score descending
tier_order = {"T1": 0, "T2": 1, "T3": 2}
catalog.sort(key=lambda c: (tier_order.get(c["priority"], 3), -c["demand_score"], c["brand"], c["model"]))

print(f"Master catalog entries: {len(catalog)}")
print(f"  T1: {sum(1 for c in catalog if c['priority'] == 'T1')}")
print(f"  T2: {sum(1 for c in catalog if c['priority'] == 'T2')}")
print(f"  T3: {sum(1 for c in catalog if c['priority'] == 'T3')}")

# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
MONEY_FMT = '$#,##0.00'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)

T1_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
T2_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
T3_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TIER_FILLS = {"T1": T1_FILL, "T2": T2_FILL, "T3": T3_FILL}


def style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER


def auto_width(ws, max_w=30):
    for col_cells in ws.columns:
        mx = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(mx + 3, max_w)


def freeze_and_filter(ws, header_row, last_col, last_row):
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = f"A{header_row + 1}"


def write_row(ws, row, values, num_cols):
    for i, v in enumerate(values, 1):
        ws.cell(row=row, column=i, value=v)
    style_data_row(ws, row, num_cols)


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = Workbook()

# ========================================================================
# TAB 1: MASTER CATALOG
# ========================================================================
ws1 = wb.active
ws1.title = "Master Catalog"

headers1 = [
    "Brand", "Model", "Size", "Width", "Aspect Ratio", "Rim Diameter",
    "Category", "Performance Tier",
    "Speed Rating", "Load Index", "Mileage Warranty",
    "Avg Retail Price (CAD)", "Min Price", "Max Price",
    "Est. Wholesale (65%)", "Our Target Price (85%)",
    "Our Margin ($)", "Our Margin (%)",
    "Demand Score (1-100)", "Top Vehicle Fitment",
    "Priority Tier", "Season Push",
]
NC1 = len(headers1)

for i, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=i, value=h)
style_header(ws1, 1, NC1)

# Column index references for formatting
COL_AVG_PRICE = 12
COL_MIN_PRICE = 13
COL_MAX_PRICE = 14
COL_WHOLESALE = 15
COL_TARGET = 16
COL_MARGIN_D = 17
COL_MARGIN_P = 18
COL_PRIORITY = 21

row = 2
for c in catalog:
    vals = [
        c["brand"], c["model"], c["size"], c["width"], c["aspect_ratio"], c["rim_diameter"],
        c["category"], c["performance_tier"],
        c["speed_rating"], c["load_index"], c["warranty_miles"],
        c["avg_price"], c["min_price"], c["max_price"],
        c["wholesale"], c["target_price"],
        c["margin_dollar"], c["margin_pct"],
        c["demand_score"], c["fitment"],
        c["priority"], c["season_push"],
    ]
    write_row(ws1, row, vals, NC1)

    # Currency formatting
    for col in [COL_AVG_PRICE, COL_MIN_PRICE, COL_MAX_PRICE, COL_WHOLESALE, COL_TARGET, COL_MARGIN_D]:
        ws1.cell(row=row, column=col).number_format = MONEY_FMT
    # Percentage formatting
    ws1.cell(row=row, column=COL_MARGIN_P).number_format = PCT_FMT
    # Priority tier coloring
    priority = c["priority"]
    if priority in TIER_FILLS:
        ws1.cell(row=row, column=COL_PRIORITY).fill = TIER_FILLS[priority]

    row += 1

last_row_1 = row - 1
freeze_and_filter(ws1, 1, NC1, last_row_1)
auto_width(ws1)

# ========================================================================
# TAB 2: PRICE INTELLIGENCE BY SIZE
# ========================================================================
ws2 = wb.create_sheet("Price Intelligence by Size")

headers2 = [
    "Size", "Width", "Aspect", "Rim",
    "# Brands", "# Models", "# Listings",
    "Min Price", "Max Price", "Avg Price", "Median Price",
    "Price Spread",
    "Top Brand 1", "Top Brand 2", "Top Brand 3",
    "Best Value Model", "Vehicle Fitment",
]
NC2 = len(headers2)

for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, NC2)

# Aggregate by size from listings
size_agg = defaultdict(lambda: {"listings": [], "brands": defaultdict(int), "models": set()})
for l in all_listings:
    sz = l.get("size", "")
    if not sz:
        continue
    sa = size_agg[sz]
    sa["listings"].append(l)
    sa["brands"][l["brand"]] += 1
    cleaned = clean_model_name(l["brand"], l.get("model", ""))
    if cleaned:
        sa["models"].add((l["brand"], cleaned))

# Also pull warranty info from catalog for best-value calc
catalog_by_size = defaultdict(list)
for c in catalog:
    if c["size"]:
        catalog_by_size[c["size"]].append(c)

row = 2
sizes_sorted = sorted(size_agg.keys(), key=lambda s: (parse_size(s)[2] or 0, parse_size(s)[0] or 0))
for sz in sizes_sorted:
    sa = size_agg[sz]
    w, a, r_d = parse_size(sz)
    prices = [l["price"] for l in sa["listings"] if l.get("price") and l["price"] > 0]
    if not prices:
        continue

    n_brands = len(sa["brands"])
    n_models = len(sa["models"])
    n_listings = len(sa["listings"])
    min_p = min(prices)
    max_p = max(prices)
    avg_p = round(statistics.mean(prices), 2)
    med_p = round(statistics.median(prices), 2)
    spread = round(max_p - min_p, 2)

    # Top 3 brands by listing count
    top_brands = sorted(sa["brands"].keys(), key=lambda b: -sa["brands"][b])[:3]
    tb1 = top_brands[0] if len(top_brands) > 0 else ""
    tb2 = top_brands[1] if len(top_brands) > 1 else ""
    tb3 = top_brands[2] if len(top_brands) > 2 else ""

    # Best value model: highest warranty / price, or lowest price if no warranty data
    best_value = ""
    cat_entries = catalog_by_size.get(sz, [])
    entries_with_warranty = [c for c in cat_entries if c.get("warranty_miles") and c.get("avg_price") and c["avg_price"] > 0]
    if entries_with_warranty:
        best = max(entries_with_warranty, key=lambda c: c["warranty_miles"] / c["avg_price"])
        best_value = f"{best['brand']} {best['model']}"
    elif cat_entries:
        priced = [c for c in cat_entries if c.get("avg_price") and c["avg_price"] > 0]
        if priced:
            best = min(priced, key=lambda c: c["avg_price"])
            best_value = f"{best['brand']} {best['model']}"

    fitment = FITMENT.get(sz, "")

    vals = [sz, w, a, r_d, n_brands, n_models, n_listings,
            min_p, max_p, avg_p, med_p, spread,
            tb1, tb2, tb3, best_value, fitment]
    write_row(ws2, row, vals, NC2)

    for col in [8, 9, 10, 11, 12]:
        ws2.cell(row=row, column=col).number_format = MONEY_FMT

    row += 1

last_row_2 = row - 1
freeze_and_filter(ws2, 1, NC2, last_row_2)
auto_width(ws2)

# ========================================================================
# TAB 3: SEASONAL CALENDAR (12 months)
# ========================================================================
ws3 = wb.create_sheet("Seasonal Calendar")

headers3 = ["Month", "Season", "Categories to Push", "Key Sizes", "Top Brands", "Action Items"]
NC3 = len(headers3)

for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, NC3)

# Derive top brands per category from catalog
def top_brands_for_category(cat, n=5):
    brand_counts = defaultdict(int)
    for c in catalog:
        if c["category"] == cat:
            brand_counts[c["brand"]] += c["listing_count"]
    return [b for b, _ in sorted(brand_counts.items(), key=lambda x: -x[1])[:n]]

winter_brands = top_brands_for_category("Winter")
as_brands = top_brands_for_category("All-Season")
at_brands = top_brands_for_category("All-Terrain")
perf_brands = top_brands_for_category("Performance")

# Top demand sizes overall
top_demand_sizes = sorted(
    size_agg.keys(),
    key=lambda s: len(size_agg[s]["listings"]),
    reverse=True,
)[:10]

# Key truck sizes
truck_sizes = [s for s in size_agg.keys() if parse_size(s)[2] and parse_size(s)[2] >= 17
               and parse_size(s)[0] and parse_size(s)[0] >= 255]

calendar_data = [
    ["January", "Deep Winter",
     "Winter",
     "225/65R17, 215/55R17, 205/55R16",
     ", ".join(winter_brands[:5]),
     "Clear remaining winter stock; promote winter packages for late buyers"],
    ["February", "Deep Winter",
     "Winter",
     "225/65R17, 265/70R17, 255/65R18",
     ", ".join(winter_brands[:5]),
     "Last-chance winter promos; begin planning spring all-season inventory"],
    ["March", "Spring Changeover",
     "All-Season, Performance",
     "225/65R17, 205/55R16, 225/45R17",
     ", ".join(as_brands[:3] + perf_brands[:2]),
     "PEAK CHANGEOVER: Push all-season swaps; spring performance campaigns"],
    ["April", "Spring Changeover",
     "All-Season, Performance",
     "215/55R17, 225/60R17, 235/55R18",
     ", ".join(as_brands[:5]),
     "Continue changeover push; promote mid-tier all-season for budget buyers"],
    ["May", "Late Spring",
     "All-Season, All-Terrain, Performance",
     "265/70R17, 245/75R16, 225/45R18",
     ", ".join(as_brands[:2] + at_brands[:2] + perf_brands[:1]),
     "Road trip prep; push all-terrain for truck owners; summer performance upsell"],
    ["June", "Summer",
     "All-Terrain, Performance",
     "265/70R17, 275/55R20, 255/70R18",
     ", ".join(at_brands[:3] + perf_brands[:2]),
     "Peak truck/SUV season; camping/towing demand; maintain all-season stock"],
    ["July", "Summer",
     "All-Terrain, All-Season",
     "275/60R20, 265/65R18, 235/70R16",
     ", ".join(at_brands[:3] + as_brands[:2]),
     "Summer road trip peak; begin early-bird winter tire pre-orders"],
    ["August", "Late Summer",
     "All-Season, Winter (pre-order)",
     "225/65R17, 215/55R17, 205/55R16",
     ", ".join(as_brands[:3] + winter_brands[:2]),
     "Back-to-school; launch winter pre-order campaigns at early-bird pricing"],
    ["September", "Fall Changeover",
     "Winter",
     "225/65R17, 205/55R16, 215/55R17",
     ", ".join(winter_brands[:5]),
     "WINTER RUSH BEGINS: Highest-margin season; push winter packages aggressively"],
    ["October", "Fall Changeover",
     "Winter",
     "225/65R17, 265/70R17, 235/65R17",
     ", ".join(winter_brands[:5]),
     "Peak winter demand; premium pricing holds; capture late-decision buyers"],
    ["November", "Early Winter",
     "Winter, All-Season (clearance)",
     "225/65R17, 215/55R17, 275/65R18",
     ", ".join(winter_brands[:5]),
     "Urgent winter buyers; clear remaining all-season overstock at discount"],
    ["December", "Winter",
     "Winter",
     "225/65R17, 255/65R18, 275/55R20",
     ", ".join(winter_brands[:5]),
     "Maintain winter stock; plan Q1 inventory; holiday promo bundles"],
]

row = 2
for cal in calendar_data:
    write_row(ws3, row, cal, NC3)
    ws3.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=10)
    row += 1

last_row_3 = row - 1
freeze_and_filter(ws3, 1, NC3, last_row_3)
auto_width(ws3, max_w=50)

# ========================================================================
# TAB 4: BRAND SUMMARY
# ========================================================================
ws4 = wb.create_sheet("Brand Summary")

headers4 = [
    "Brand", "# Models", "# Sizes", "Avg Price", "Min Price", "Max Price",
    "Market Share %", "Top Category",
    "Has Winter?", "Has All-Terrain?",
    "Min Warranty", "Max Warranty",
]
NC4 = len(headers4)

for i, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, NC4)

# Aggregate by brand
brand_agg = defaultdict(lambda: {
    "models": set(), "sizes": set(), "prices": [], "listings": 0,
    "categories": defaultdict(int), "warranties": [],
})
for c in catalog:
    b = brand_agg[c["brand"]]
    b["models"].add(c["model"])
    b["sizes"].add(c["size"])
    if c["avg_price"]:
        b["prices"].append(c["avg_price"])
    b["listings"] += c["listing_count"]
    b["categories"][c["category"]] += c["listing_count"]
    if c.get("warranty_miles"):
        b["warranties"].append(c["warranty_miles"])

total_listings_all = sum(b["listings"] for b in brand_agg.values())
brands_sorted = sorted(brand_agg.keys(), key=lambda b: -brand_agg[b]["listings"])

row = 2
for brand in brands_sorted:
    b = brand_agg[brand]
    n_models = len(b["models"])
    n_sizes = len(b["sizes"])
    avg_p = round(statistics.mean(b["prices"]), 2) if b["prices"] else None
    min_p = round(min(b["prices"]), 2) if b["prices"] else None
    max_p = round(max(b["prices"]), 2) if b["prices"] else None
    share = round(b["listings"] / total_listings_all * 100, 1) if total_listings_all else 0
    top_cat = max(b["categories"].keys(), key=lambda k: b["categories"][k]) if b["categories"] else ""
    has_winter = "Yes" if b["categories"].get("Winter", 0) > 0 else "No"
    has_at = "Yes" if b["categories"].get("All-Terrain", 0) > 0 else "No"
    min_war = min(b["warranties"]) if b["warranties"] else None
    max_war = max(b["warranties"]) if b["warranties"] else None

    vals = [brand, n_models, n_sizes, avg_p, min_p, max_p,
            share, top_cat, has_winter, has_at, min_war, max_war]
    write_row(ws4, row, vals, NC4)

    for col in [4, 5, 6]:
        ws4.cell(row=row, column=col).number_format = MONEY_FMT

    row += 1

last_row_4 = row - 1
freeze_and_filter(ws4, 1, NC4, last_row_4)
auto_width(ws4)

# ========================================================================
# TAB 5: TOP MODELS BY DEMAND (top 50)
# ========================================================================
ws5 = wb.create_sheet("Top Models by Demand")

headers5 = [
    "Rank", "Brand", "Model", "Category",
    "# Sizes", "Demand Score",
    "Speed Rating", "Load Index", "Mileage Warranty",
    "Avg Price", "Min Price", "Max Price",
    "Est. Wholesale", "Our Target", "Our Margin ($)", "Our Margin (%)",
    "Performance Tier", "Top Vehicle Fitments",
]
NC5 = len(headers5)

for i, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, NC5)

# Aggregate by (brand, model) across sizes
model_agg = defaultdict(lambda: {
    "sizes": set(), "prices": [], "demand_scores": [], "listing_count": 0,
    "category": "", "speed_rating": "", "load_index": "", "warranty_miles": None,
    "fitments": set(),
})
for c in catalog:
    key = (c["brand"], c["model"])
    ma = model_agg[key]
    ma["sizes"].add(c["size"])
    if c["avg_price"]:
        ma["prices"].append(c["avg_price"])
    ma["demand_scores"].append(c["demand_score"])
    ma["listing_count"] += c["listing_count"]
    if c["category"]:
        ma["category"] = c["category"]
    if c["speed_rating"]:
        ma["speed_rating"] = c["speed_rating"]
    if c["load_index"]:
        ma["load_index"] = c["load_index"]
    if c.get("warranty_miles"):
        ma["warranty_miles"] = c["warranty_miles"]
    if c["fitment"]:
        ma["fitments"].add(c["fitment"])

# Rank by max demand score, then by listing count — exclude "Unspecified"
models_ranked = sorted(
    [k for k in model_agg.keys() if k[1] != "Unspecified"],
    key=lambda k: (-max(model_agg[k]["demand_scores"]), -model_agg[k]["listing_count"]),
)

row = 2
for rank, key in enumerate(models_ranked[:50], 1):
    brand, model = key
    ma = model_agg[key]
    demand = max(ma["demand_scores"])
    avg_p = round(statistics.mean(ma["prices"]), 2) if ma["prices"] else None
    min_p = round(min(ma["prices"]), 2) if ma["prices"] else None
    max_p = round(max(ma["prices"]), 2) if ma["prices"] else None
    wholesale = round(avg_p * 0.65, 2) if avg_p else None
    target = round(avg_p * 0.85, 2) if avg_p else None
    margin_d = round(target - wholesale, 2) if (target and wholesale) else None
    margin_p = round(margin_d / target, 4) if (target and margin_d and target > 0) else None
    tier = performance_tier(avg_p)
    fitments = "; ".join(sorted(ma["fitments"])[:3])

    vals = [rank, brand, model, ma["category"],
            len(ma["sizes"]), demand,
            ma["speed_rating"], ma["load_index"], ma["warranty_miles"],
            avg_p, min_p, max_p,
            wholesale, target, margin_d, margin_p,
            tier, fitments]
    write_row(ws5, row, vals, NC5)

    for col in [10, 11, 12, 13, 14, 15]:
        ws5.cell(row=row, column=col).number_format = MONEY_FMT
    ws5.cell(row=row, column=16).number_format = PCT_FMT

    row += 1

last_row_5 = row - 1
freeze_and_filter(ws5, 1, NC5, last_row_5)
auto_width(ws5)

# ========================================================================
# TAB 6: ALL LISTINGS (RAW)
# ========================================================================
ws6 = wb.create_sheet("All Listings (Raw)")

headers6 = [
    "Brand", "Model (Raw)", "Model (Cleaned)", "Price (CAD)",
    "Size", "Type", "Retailer", "Source",
]
NC6 = len(headers6)

for i, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=i, value=h)
style_header(ws6, 1, NC6)

row = 2
for l in sorted(all_listings, key=lambda x: (x.get("size", ""), x.get("brand", ""), x.get("price", 0))):
    raw_model = l.get("model", "")
    cleaned = clean_model_name(l["brand"], raw_model)

    vals = [
        l.get("brand", ""),
        raw_model,
        cleaned,
        l.get("price", 0),
        l.get("size", ""),
        l.get("type", "unknown"),
        l.get("retailer", "unknown"),
        l.get("source", ""),
    ]
    write_row(ws6, row, vals, NC6)
    ws6.cell(row=row, column=4).number_format = MONEY_FMT

    row += 1

last_row_6 = row - 1
freeze_and_filter(ws6, 1, NC6, last_row_6)
auto_width(ws6)

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------
wb.save(str(OUTPUT_PATH))

# ---------------------------------------------------------------------------
# Summary stats
# ---------------------------------------------------------------------------
print(f"\n{'='*60}")
print(f"  SAVED: {OUTPUT_PATH}")
print(f"{'='*60}")
print(f"  Tab 1 — Master Catalog:           {last_row_1 - 1} rows")
print(f"  Tab 2 — Price Intelligence:        {last_row_2 - 1} sizes")
print(f"  Tab 3 — Seasonal Calendar:         {last_row_3 - 1} months")
print(f"  Tab 4 — Brand Summary:             {last_row_4 - 1} brands")
print(f"  Tab 5 — Top Models by Demand:      {min(50, last_row_5 - 1)} models")
print(f"  Tab 6 — All Listings (Raw):        {last_row_6 - 1} listings")
print(f"{'='*60}")
print(f"  Unique brands:  {len(brands_sorted)}")
print(f"  Unique sizes:   {len(size_agg)}")
print(f"  Total listings: {len(all_listings)}")
print(f"  Spec matches:   {len(matched_spec_keys)}")
t1 = sum(1 for c in catalog if c["priority"] == "T1")
t2 = sum(1 for c in catalog if c["priority"] == "T2")
t3 = sum(1 for c in catalog if c["priority"] == "T3")
print(f"  Priority: T1={t1}  T2={t2}  T3={t3}")
print(f"{'='*60}")
